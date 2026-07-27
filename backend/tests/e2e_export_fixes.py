"""Focused E2E for the three video-feedback fixes:
  A) buyer My Results LOST lines show the MODEL quantity, not 1
  B) Export Center simplified (backend unchanged for report-pack, verified it still builds)
  C) Bid comparison export is fast, has a Winner column, and highlights the winning cell.
Seeds an in-memory DB directly (no Postgres) and calls the real service/route functions.
"""
import io, time, sys
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from openpyxl import load_workbook

from app.db.base import Base
from app.models.user import User
from app.models.bid_round import BidRound
from app.models.master_item import MasterItem
from app.models.bid_file import BidFile
from app.models.bid_line import BidLine
from app.models.deal import Deal
from app.core.security import hash_password
from app.services.export_service import export_bid_comparison_excel, export_report_pack_zip

engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
Base.metadata.create_all(bind=engine)
db = sessionmaker(bind=engine)()

fails = []
def check(cond, msg):
    print(("PASS" if cond else "FAIL"), "-", msg)
    if not cond:
        fails.append(msg)

# ── Seed: 1 round, 3 buyers, 3 models (qty 12/8/58), each buyer bids each model ──
rnd = BidRound(name="Round 49", status="complete", commodity="memory")
db.add(rnd); db.commit(); db.refresh(rnd)

buyers = []
for i, (co) in enumerate(["Alpha Traders", "Beta Recyclers", "Gamma Wholesale"]):
    u = User(email=f"b{i}@x.com", hashed_password=hash_password("x"),
             full_name=f"Buyer {i}", company_name=co, role="buyer", is_active=True)
    db.add(u); buyers.append(u)
db.commit()
for u in buyers:
    db.refresh(u)

models = []
for pn, qty in [("15-13079-02", 12), ("01DE973", 8), ("7X77A01303", 58)]:
    m = MasterItem(bid_round_id=rnd.id, part_number=pn, part_number_normalized=pn.lower(),
                   description=f"desc {pn}", quantity=qty,
                   unit_details=[{"Serial": f"S{pn}{k}", "UID": f"U{pn}{k}"} for k in range(qty)])
    db.add(m); models.append(m)
db.commit()
for m in models:
    db.refresh(m)

# Prices: higher = better (reverse auction where seller wants max). Buyer 0 wins model 0 & 1,
# buyer 2 wins model 2. Buyer 1 loses everything (this is the loser whose LOST qty we check).
price_grid = {
    # model_idx: {buyer_idx: unit_price}
    0: {0: 100.0, 1: 90.0, 2: 95.0},   # winner buyer0
    1: {0: 50.0,  1: 40.0, 2: 45.0},   # winner buyer0
    2: {0: 20.0,  1: 18.0, 2: 25.0},   # winner buyer2
}
winner_of = {0: 0, 1: 0, 2: 2}

bfiles = {}
for bi, u in enumerate(buyers):
    f = BidFile(bid_round_id=rnd.id, buyer_id=u.id, filename=f"buyer{bi}.xlsx", file_path=f"/tmp/buyer{bi}.xlsx", status="processed")
    db.add(f); db.commit(); db.refresh(f)
    bfiles[bi] = f

for mi, m in enumerate(models):
    for bi, u in enumerate(buyers):
        price = price_grid[mi][bi]
        is_win = (winner_of[mi] == bi)
        line = BidLine(
            bid_file_id=bfiles[bi].id, bid_round_id=rnd.id, buyer_id=u.id, master_item_id=m.id,
            raw_part_number=m.part_number, unit_price=price,
            quantity=m.quantity,  # _assign_match sets this to master.quantity in real flow
            total_price=price * m.quantity, match_method="exact", match_score=100.0,
            match_status="matched", is_winner=is_win,
            fluffed_loss_price=None if is_win else round(price_grid[mi][winner_of[mi]] * 1.03, 2),
        )
        db.add(line)
db.commit()

# Deals for winners (authoritative)
for mi, m in enumerate(models):
    bi = winner_of[mi]
    wline = db.query(BidLine).filter(BidLine.master_item_id == m.id, BidLine.buyer_id == buyers[bi].id).first()
    d = Deal(bid_round_id=rnd.id, master_item_id=m.id, winning_buyer_id=buyers[bi].id,
             winning_bid_line_id=wline.id, part_number=m.part_number, description=m.description,
             quantity=m.quantity, winning_price=wline.unit_price,
             total_value=wline.unit_price * m.quantity, status="approved")
    db.add(d)
db.commit()

# ── FIX A: buyer My Results LOST quantity == model quantity, not 1 ──
from app.api.routes import buyer as buyer_routes
# Simulate the loser (buyer1) — call the underlying function with a stub dependency
res = buyer_routes.my_results_for_round.__wrapped__ if hasattr(buyer_routes.my_results_for_round, "__wrapped__") else buyer_routes.my_results_for_round
out = res(rnd.id, db=db, buyer=buyers[1])
lost = [r for r in out["results"] if r["outcome"] == "LOST"]
check(len(lost) == 3, f"loser has 3 LOST lines (got {len(lost)})")
check(all(r["quantity"] in (12, 8, 58) for r in lost),
      f"LOST lines carry model qty not 1: {[r['quantity'] for r in lost]}")
check(not any(r["quantity"] == 1 for r in lost), "no LOST line shows qty 1")
# rollup lost counts should sum model quantities (12+8+58 = 78)
total_lost_units = sum(g["lost"] for g in out["rollup"])
check(total_lost_units == 78, f"rollup lost units = 78 (got {total_lost_units})")

# ── FIX C: comparison export fast + Winner column + winning cell highlight ──
t0 = time.time()
data = export_bid_comparison_excel(db, rnd.id)
elapsed = time.time() - t0
check(elapsed < 2.0, f"comparison export fast (<2s), took {elapsed:.2f}s")
wb = load_workbook(io.BytesIO(data))
ws = wb["Bid Comparison"]
hdr = [c.value for c in ws[1]]
check("Winner" in hdr, f"has Winner column (headers: {hdr})")
check("Winning Price" in hdr, "has Winning Price column")
# each buyer company is a column
for co in ["Alpha Traders", "Beta Recyclers", "Gamma Wholesale"]:
    check(co in hdr, f"buyer column present: {co}")
check(ws.max_row == 4, f"one row per model = 3 data rows (max_row={ws.max_row})")
# Winner column values should be company names
winner_col_idx = hdr.index("Winner") + 1
winners_seen = {ws.cell(row=r, column=winner_col_idx).value for r in range(2, ws.max_row + 1)}
check("Alpha Traders" in winners_seen and "Gamma Wholesale" in winners_seen,
      f"winner companies shown in Winner column: {winners_seen}")
# a winning cell is highlighted green (C6EFCE)
green_cells = 0
for r in range(2, ws.max_row + 1):
    for c in range(4, 4 + 3):
        cell = ws.cell(row=r, column=c)
        if cell.fill and cell.fill.fgColor and str(cell.fill.fgColor.rgb).endswith("C6EFCE"):
            green_cells += 1
check(green_cells == 3, f"3 winning cells highlighted green (got {green_cells})")

# ── FIX B sanity: report pack still builds and contains the comparison + razor ──
pack = export_report_pack_zip(db, rnd.id)
import zipfile
zf = zipfile.ZipFile(io.BytesIO(pack))
names = zf.namelist()
check(any("bid_comparison" in n for n in names), f"report pack has comparison: {names}")
check(any("razor_per_customer" in n for n in names), "report pack has razor per-customer")
check(any("award_sheets" in n for n in names), "report pack has award sheets")
check("_errors.txt" not in names, f"report pack built with no errors: {names}")

print()
if fails:
    print(f"❌ {len(fails)} FAILED:")
    for f in fails:
        print("  -", f)
    sys.exit(1)
print("✅ ALL EXPORT-FIX CHECKS PASSED")
