"""
Performance regression guardrails — each test encodes a real, measured incident so the
exact same regression fails CI instead of reaching a user.

Thresholds are set well above normal operating numbers (measured inline in each test's
docstring) so the tests stay reliable in CI without becoming flaky, while still catching a
real regression by an order of magnitude.
"""
import time
import asyncio
import io

import openpyxl
import pytest
from sqlalchemy import event

from app.core.security import verify_password, hash_password
from app.core.executors import file_parsing_executor
from app.services.file_parser import parse_master_file
from app.models.bid_round import BidRound
from app.models.bid_file import BidFile
from app.models.bid_line import BidLine
from app.models.master_item import MasterItem
from app.models.deal import Deal
from app.models.user import User
from app.services.export_service import export_disposition_report


# ── Login latency under file-upload contention ─────────────────────────────────
# Incident: asyncio.to_thread() for file parsing shared the same default executor
# login's bcrypt check uses. Measured: 91ms -> 13,581ms under 2 concurrent uploads on a
# small pool. Fix: a dedicated executor (app/core/executors.py) for parsing only.

def _make_xlsx(num_rows: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Part Number", "Description", "Quantity", "Reserve Price"])
    for i in range(num_rows):
        ws.append([f"PN-{i:05d}", f"Test Item {i}", 1, 10.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_login_latency_unaffected_by_concurrent_file_parsing():
    """Login's bcrypt check must stay fast even while the dedicated file-parsing
    executor is fully saturated by large uploads. Threshold: 2 seconds — normal
    bcrypt latency is under 200ms; the incident this guards against took 13.5s."""
    loop = asyncio.new_event_loop()
    try:
        pw_hash = hash_password("guardrail-test-password")
        big_file = _make_xlsx(2000)

        async def run():
            async def fake_upload():
                await loop.run_in_executor(file_parsing_executor, parse_master_file, big_file, "big.xlsx")

            # Saturate the dedicated parsing pool with concurrent heavy uploads.
            upload_tasks = [asyncio.ensure_future(fake_upload()) for _ in range(3)]
            await asyncio.sleep(0.1)  # let uploads grab the pool's worker threads

            t0 = time.time()
            await asyncio.to_thread(verify_password, "guardrail-test-password", pw_hash)
            login_latency = time.time() - t0

            await asyncio.gather(*upload_tasks)
            return login_latency

        latency = loop.run_until_complete(run())
        assert latency < 2.0, (
            f"Login took {latency:.2f}s while 3 file uploads were parsing concurrently — "
            "this is the thread-pool-contention regression (was measured at 13.5s before "
            "the dedicated executor fix). Check that file parsing still uses "
            "app.core.executors.file_parsing_executor, not asyncio's default executor."
        )
    finally:
        loop.close()


# ── File parsing must never lose or duplicate rows ──────────────────────────────
# Incident: a hidden "rollup" sheet alongside per-tab breakdowns doubled every row;
# unit-level aggregation silently merged distinct units. Both fixed to be exact 1:1.

@pytest.mark.parametrize("num_rows", [7, 100, 537])
def test_file_parsing_preserves_exact_row_count(num_rows):
    content = _make_xlsx(num_rows)
    rows = parse_master_file(content, "guardrail_test.xlsx")
    assert len(rows) == num_rows, (
        f"Uploaded {num_rows} rows but parser returned {len(rows)} — file parsing must "
        "produce an exact 1:1 row count with no silent loss or duplication."
    )
    part_numbers = [r["part_number"] for r in rows]
    assert len(part_numbers) == len(set(part_numbers)), "Duplicate part numbers in parsed output."


# ── Upload consolidates by MODEL, preserving every spec column and every device ─────
# Requirement (ThinkTLS, 2026-07-23): a unit-level file is bid at the MODEL level — "qty (8)
# 15-13079-02", not 8 separate lines. Uploading must (a) sum the quantity per model, (b) keep
# every model-level spec column, and (c) preserve EVERY physical device's Serial/UID in
# unit_details so the per-winner Razor output can expand a won model back to its devices. The
# per-device identifier columns move to unit_details and don't appear as model-level spec cols.

def _make_unit_level_xlsx(num_rows: int, spec_cols: list[str], blank_col: str | None = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    # Serial makes it a unit-level file; Model repeats so rows consolidate by model.
    cols = ["Model", "Serial", "MFG"] + spec_cols
    ws.append(cols)
    for i in range(num_rows):
        vals = []
        for c in spec_cols:
            # blank_col is empty on the FIRST rows and only filled later — the real "Drive bays"
            # shape that used to shove a column to the end of the template.
            if c == blank_col:
                vals.append("" if i < 3 else f"late-{i}")
            else:
                vals.append(f"{c}-val")   # model-level: same across a model's devices
        ws.append([f"MODEL-{i%40}", f"SN{i:06d}", "DELL"] + vals)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_upload_consolidates_by_model_and_preserves_devices():
    spec_cols = [f"Spec Column {chr(65+i)}" for i in range(12)]
    num_rows = 750  # 40 distinct models (MODEL-0..39)
    content = _make_unit_level_xlsx(num_rows, spec_cols)
    rows = parse_master_file(content, "big_unit_file.xlsx")

    # Consolidated to one row per model, with the device count summed and every device kept.
    assert len(rows) == 40, f"Expected 40 models, got {len(rows)} — file must consolidate by model."
    assert sum(r["quantity"] for r in rows) == num_rows, "Total device quantity must equal input rows."
    assert sum(len(r.get("unit_details") or []) for r in rows) == num_rows, (
        "Every physical device's Serial/UID must be preserved in unit_details for the Razor output."
    )
    for r in rows:
        preserved = set(r.get("extra_columns") or {})
        missing = set(spec_cols) - preserved
        assert not missing, f"Spec columns dropped: {sorted(missing)} — nothing may be missed."
        assert r["manufacturer"] == "DELL"
        # Serial is a per-device identifier — it belongs in unit_details, not as a model spec col.
        assert "Serial" not in preserved
        assert all("Serial" in d for d in (r.get("unit_details") or []))


# ── Model-level spec columns keep the source file's order ────────────────────────
# extra_columns must round-trip in the file's column order (it was physically jsonb, which
# discards key order; and blank cells were dropped per row, reordering columns). Identifier
# columns (Serial/UID) are excluded now that they live in unit_details.

def test_extra_columns_keep_source_order_including_blank_columns():
    spec_cols = ["Alpha Spec", "Blank Until Later", "Zeta Spec", "Mid Spec"]
    content = _make_unit_level_xlsx(20, spec_cols, blank_col="Blank Until Later")
    rows = parse_master_file(content, "order_test.xlsx")

    # Only the non-standard spec columns remain, in the file's order: Model→part_number,
    # MFG→manufacturer and Serial→unit_details are all surfaced elsewhere, not repeated here.
    expected = spec_cols
    for r in rows:
        got = list((r.get("extra_columns") or {}).keys())
        assert got == expected, (
            f"extra_columns order drifted from the source file.\n expected: {expected}\n got:      {got}"
        )
    # The blank column must still be present (as "") in order, not dropped.
    first = rows[0]["extra_columns"]
    assert "Blank Until Later" in first and first["Blank Until Later"] == ""
    assert rows[-1]["extra_columns"]["Blank Until Later"].startswith("late-")


def test_buyer_price_column_not_duplicated_as_spec_column():
    """An admin file's own price box ("Offer") must not survive as a locked spec column —
    the template already renders it as the editable "Unit Price ($)". Keeping both put a dead
    duplicate next to the real one."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Offer", "Model", "Serial", "MFG", "CPU Type"])
    for i in range(10):
        ws.append([None, f"MODEL-{i}", f"SN{i:04d}", "DELL", "i7"])
    buf = io.BytesIO()
    wb.save(buf)

    rows = parse_master_file(buf.getvalue(), "offer_col.xlsx")
    for r in rows:
        keys = {k.lower() for k in (r.get("extra_columns") or {})}
        assert "offer" not in keys, "Buyer price placeholder 'Offer' leaked in as a spec column"
        assert "CPU Type" in (r.get("extra_columns") or {})


# ── Side-by-side pivot bid tables must not double-count section 0 ────────────────
# Incident: the ThinkTLS memory "Bid Tables" sheet holds several pivot bid tables laid out
# side-by-side, separated by blank spacer columns (e.g. "64GB DDR4", "32GB DDR4", ...).
# The parser split it into per-section blocks AND kept a flattened whole-sheet parse of the
# same sheet; flattening a multi-section pivot only captures the first section (suffixing the
# rest .1/.2/…), and merging that flattened copy back with the blocks DOUBLED the first
# section's quantities and injected junk "Count of Model" spec columns into the template.

def _make_pivot_xlsx() -> bytes:
    """Two pivot bid tables side-by-side, separated by a blank spacer column — mirrors the
    real ThinkTLS memory 'Bid Tables' layout. Each block: Row Labels | Count of Model | Qty |
    Bid Unit | Bid Ext. Section A quantities are distinctive so any doubling is unmissable."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid Tables"
    hdr = ["Row Labels", "Count of Model", "Qty", "Bid Unit", "Bid Ext"]
    # header row (col 0-4 = section A, col 5 = spacer, col 6-10 = section B)
    ws.append(hdr + [None] + hdr)
    section_a = [("A-100", 3), ("A-200", 111), ("A-300", 7)]
    section_b = [("B-100", 5), ("B-200", 9)]
    for i in range(max(len(section_a), len(section_b))):
        row = []
        if i < len(section_a):
            pn, q = section_a[i]; row += [pn, q, q, None, 0]
        else:
            row += [None] * 5
        row += [None]  # spacer
        if i < len(section_b):
            pn, q = section_b[i]; row += [pn, q, q, None, 0]
        else:
            row += [None] * 5
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_side_by_side_pivot_not_double_counted():
    rows = parse_master_file(_make_pivot_xlsx(), "memory_bid_tables.xlsx")
    by_pn = {r["part_number"]: r for r in rows}
    # All part numbers from BOTH sections present exactly once.
    assert set(by_pn) == {"A-100", "A-200", "A-300", "B-100", "B-200"}, (
        f"Pivot sections not all captured exactly once: got {sorted(by_pn)}"
    )
    # Section-0 quantity must be its real value, not doubled (111, never 222).
    assert by_pn["A-200"]["quantity"] == 111, (
        f"A-200 qty={by_pn['A-200']['quantity']} — the flattened-vs-pivot double-count "
        "regression is back (expected 111, doubling would give 222)."
    )
    # No pivot-count junk column leaked into the template spec columns.
    for r in rows:
        for k in (r.get("extra_columns") or {}):
            assert "count of" not in k.lower(), f"Pivot noise column '{k}' leaked into extra_columns"


# ── Matching must stay fast at real round size ──────────────────────────────────
# Incident: the fuzzy tier looped over every master item in Python for every bid line that
# missed the exact-match index. Once drive/memory rounds became per-unit, a memory round is
# 9,305 masters x 2 buyers = 18,610 lines — 173M comparisons. Measured 8.6ms/line, ~161s on a
# dev box (many minutes on the hosted tier), and the UI just sat at "0%" looking frozen.
# Fixed with rapidfuzz process.extractOne (C++ scan + score_cutoff): ~0.9ms/line.

def test_fuzzy_matching_stays_fast_on_large_rounds():
    from app.services.matcher import match_bid_lines
    from app.models.bid_line import BidLine as _BidLine
    from app.models.master_item import MasterItem as _MasterItem

    masters = []
    for i in range(1, 9306):
        m = _MasterItem()
        m.id = i
        m.part_number = f"PN{i}"
        m.part_number_normalized = f"HMA84GR7AFR4NUH{i:05d}"
        m.quantity = 1
        masters.append(m)

    # Worst case: none of these hit the exact-match index, so every one goes through fuzzy.
    lines = []
    for i in range(200):
        l = _BidLine()
        l.buyer_id = 1
        l.raw_part_number = f"NOMATCH-{i}"
        l.normalized_part_number = f"NOMATCHXYZ{i:05d}"
        l.quantity = 1
        lines.append(l)

    t0 = time.time()
    match_bid_lines(lines, masters)
    per_line_ms = (time.time() - t0) / len(lines) * 1000

    # The Python-loop version measured ~8.6ms/line. 4ms is a wide margin over the ~0.9ms the
    # rapidfuzz path costs, while still failing loudly if the O(N*M) loop ever comes back.
    assert per_line_ms < 4.0, (
        f"Fuzzy matching is {per_line_ms:.1f}ms per line against 9,305 masters — the "
        f"per-line Python scan over every master item is back. At a real memory round "
        f"(18,610 lines) this costs {per_line_ms * 18610 / 1000:.0f}s and the round appears "
        f"frozen at 0%. Matching must use rapidfuzz process.extractOne."
    )
    assert all(l.match_status == "exception" for l in lines)


# ── Template cache must invalidate when item CONTENT changes ─────────────────────
# Incident: the bid-template cache was keyed on row COUNT only. When an admin re-uploaded an
# edited master file with the same number of rows (fixed a description / quantity / spec),
# download_template served the STALE cached template, so buyers downloaded the old data.
# Fix: cache key is now a content fingerprint of every rendered field.

def test_template_cache_invalidates_on_content_edit():
    from types import SimpleNamespace
    from app.services import template_generator as tg

    def _item(desc):
        return SimpleNamespace(
            id=1, bid_round_id=1, row_number=1, part_number="PN1", part_number_normalized="PN1",
            description=desc, manufacturer="DELL", quantity=5, reserve_price=None,
            category="", extra_columns=None,
        )

    class _Q:
        def __init__(self, res): self.res = res
        def filter(self, *a, **k): return self
        def order_by(self, *a, **k): return self
        def all(self): return self.res if isinstance(self.res, list) else []
        def first(self): return self.res if not isinstance(self.res, list) else None

    class _DB:
        def __init__(self, rnd, items): self.rnd, self.items = rnd, items
        def query(self, model):
            return _Q(self.rnd if "BidRound" in getattr(model, "__name__", "") else self.items)

    rnd = SimpleNamespace(id=1, name="R", commodity="c", customer="x", submission_deadline=None)
    tg._template_cache.clear()

    def _desc_in_template(tpl_bytes):
        ws = openpyxl.load_workbook(io.BytesIO(tpl_bytes))["Bid Template"]
        headers = [c.value for c in ws[1]]
        return ws.cell(row=2, column=headers.index("Description") + 1).value

    first = tg.generate_bid_template(_DB(rnd, [_item("Original Widget")]), 1)
    assert _desc_in_template(first) == "Original Widget"

    # Same row count, edited description — must NOT return the stale cached bytes.
    edited = tg.generate_bid_template(_DB(rnd, [_item("CORRECTED Widget")]), 1)
    assert _desc_in_template(edited) == "CORRECTED Widget", (
        "Stale-template-cache regression: an edited master with the same row count still "
        "served the old cached template. Cache must key on content, not row count."
    )
    # Identical content should still hit the cache (byte-identical output).
    again = tg.generate_bid_template(_DB(rnd, [_item("CORRECTED Widget")]), 1)
    assert again == edited


# ── Export N+1 query regression ─────────────────────────────────────────────────
# Incident: export_disposition_report queried Deal + BidLine per master item (3 queries
# x N). At 1818 real master items this measured 14.0s. Fixed via batch-fetch up front.

def test_disposition_export_has_no_n_plus_one(db):
    """500 master items + 500 deals must export in well under 1 second if batch-fetched
    correctly. The N+1 version of this code measured 14s at 1818 items (~3.8x this size) —
    a query-per-row regression would push this test from milliseconds to multiple seconds."""
    round_ = BidRound(name="Perf Guardrail Round", commodity="test", status="complete", master_file_uploaded=True)
    db.add(round_)
    db.flush()

    buyer = User(
        email="perfguard@test.com", hashed_password=hash_password("pass"),
        full_name="Perf Guard Buyer", role="buyer", is_active=True,
    )
    db.add(buyer)
    db.flush()

    bid_file = BidFile(
        bid_round_id=round_.id, buyer_id=buyer.id,
        filename="perfguard.xlsx", file_path="/tmp/perfguard.xlsx",
    )
    db.add(bid_file)
    db.flush()

    for i in range(500):
        m = MasterItem(
            bid_round_id=round_.id, part_number=f"PG-{i:04d}", part_number_normalized=f"PG{i:04d}",
            description=f"Item {i}", quantity=1, reserve_price=10.0,
        )
        db.add(m)
        db.flush()
        line = BidLine(
            bid_file_id=bid_file.id, bid_round_id=round_.id, buyer_id=buyer.id,
            master_item_id=m.id, raw_part_number=m.part_number, unit_price=15.0,
            quantity=1, match_status="matched", is_winner=True,
        )
        db.add(line)
        db.flush()
        db.add(Deal(
            bid_round_id=round_.id, master_item_id=m.id, winning_buyer_id=buyer.id,
            winning_bid_line_id=line.id, part_number=m.part_number, description=m.description,
            quantity=1, winning_price=15.0, total_value=15.0, status="approved",
        ))
    db.commit()

    # Count actual SQL statements rather than timing wall-clock: the test DB is in-memory
    # SQLite with near-zero per-query latency, so an N+1 regression here would still
    # complete in milliseconds and a timing-only assertion would never catch it — the real
    # 14s incident only showed up over a real network connection to Postgres. Counting
    # statements detects the O(N) vs O(1) query pattern directly, regardless of DB backend.
    statement_count = 0

    def _count_statements(*args, **kwargs):
        nonlocal statement_count
        statement_count += 1

    event.listen(db.bind, "before_cursor_execute", _count_statements)
    try:
        t0 = time.time()
        data = export_disposition_report(db, round_.id)
        elapsed = time.time() - t0
    finally:
        event.remove(db.bind, "before_cursor_execute", _count_statements)

    assert statement_count < 20, (
        f"Disposition export of 500 master items ran {statement_count} SQL statements — "
        "expected a small constant number regardless of row count. This is the N+1 query "
        "regression (3 queries per master item x 500 = 1500+ before the fix, measured "
        "as 14s against real Postgres at 1818 items). Check export_disposition_report "
        "still batch-fetches deals/bid-lines/buyers before the per-row loop."
    )
    assert len(data) > 0
    assert elapsed < 5.0  # sanity bound even against the slower in-memory SQLite test DB
