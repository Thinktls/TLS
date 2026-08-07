"""Verify the 2026-07-30 feedback batch:
  3) part# UPPERCASE, descriptions proper case
  4) results email shows WON items (not outbid)
  5) award sheet prices in USD number format
  6) brokers@thinktls.com in buyer-facing email
  7) PO instruction to winners (email + award sheet)
"""
import io, sys
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from openpyxl import load_workbook

from app.db.base import Base
from app.models.user import User
from app.models.bid_round import BidRound
from app.models.master_item import MasterItem
from app.models.bid_file import BidFile
from app.models.bid_line import BidLine
from app.models.deal import Deal
from app.core.security import hash_password
from app.services.normalizer import format_part_number, normalize_description
from app.services.email_templates import results_email, BROKERS_EMAIL
from app.services.export_service import export_buyer_award_sheet

fails = []
def check(cond, msg):
    print(("PASS" if cond else "FAIL"), "-", msg)
    if not cond:
        fails.append(msg)

# ── 3) formatting helpers ──
check(format_part_number("h0h72108clar8000") == "H0H72108CLAR8000", "part# uppercased")
check(normalize_description("hgst hus724040als640 4tb sas hdd") == "HGST HUS724040ALS640 4TB SAS HDD",
      "description proper-cased: " + normalize_description("hgst hus724040als640 4tb sas hdd"))
check(normalize_description("intel intel ssdsc2kb240g8") == "Intel Intel SSDSC2KB240G8", "brand + pn token")

# ── 4/6/7) results email shows WON items + brokers PO ──
won_items = [
    {"part_number": "H0H72108CLAR8000", "description": "Hitachi H0H72108CLAR8000", "quantity": 55, "your_price": 136.62},
    {"part_number": "EG002400JWJNN", "description": "HPE 2.4TB SAS", "quantity": 5, "your_price": 80.0},
]
lost_items = [
    {"part_number": "KPM5XRUG1T92", "description": "toshiba kpm5xrug1t92", "quantity": 12,
     "your_price": 150.0, "winning_price": 241.36},
]
subject, html = results_email("David Buyer", "Drives - Test", won_count=2, lost_count=178,
                              portal_url="http://x/portal/results?round=52",
                              won_items=won_items, lost_items=lost_items)
check("Items You Won" in html, "email has 'Items You Won' section")
check("Items You Were Outbid" in html, "email ALSO shows the full outbid results")
check("KPM5XRUG1T92" in html and "$241.36" in html, "lost item + winning price rendered")
check(BROKERS_EMAIL in html and "brokers@thinktls.com" in html, "email has brokers@thinktls.com")
check("issue a PO" in html and "24 hours" in html, "email has PO-within-24h instruction")
check("H0H72108CLAR8000" in html and "$136.62" in html, "won item + price rendered")
check("55" in html, "won qty rendered")

# ── 5/7) award sheet USD number format + PO note + model qty ──
engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
Base.metadata.create_all(bind=engine)
db = sessionmaker(bind=engine)()

rnd = BidRound(name="Drives - Test", status="complete", commodity="drives")
db.add(rnd); db.commit(); db.refresh(rnd)
buyer = User(email="d@x.com", hashed_password=hash_password("x"), full_name="David", company_name="Dbuy", role="buyer", is_active=True)
db.add(buyer); db.commit(); db.refresh(buyer)
# won model (qty 55) + lost model (qty 20)
mw = MasterItem(bid_round_id=rnd.id, part_number="H0H72108CLAR8000", part_number_normalized="h0h72108clar8000",
                description="hitachi h0h72108clar8000", quantity=55)
ml = MasterItem(bid_round_id=rnd.id, part_number="832414-B21", part_number_normalized="832414b21",
                description="hpe 832414-b21 480gb sata ssd", quantity=20)
db.add_all([mw, ml]); db.commit(); [db.refresh(m) for m in (mw, ml)]
bf = BidFile(bid_round_id=rnd.id, buyer_id=buyer.id, filename="d.xlsx", file_path="/tmp/d.xlsx", status="processed")
db.add(bf); db.commit(); db.refresh(bf)
lw = BidLine(bid_file_id=bf.id, bid_round_id=rnd.id, buyer_id=buyer.id, master_item_id=mw.id,
             raw_part_number="H0H72108CLAR8000", unit_price=136.62, quantity=55, total_price=136.62*55,
             match_method="exact", match_status="matched", is_winner=True)
ll = BidLine(bid_file_id=bf.id, bid_round_id=rnd.id, buyer_id=buyer.id, master_item_id=ml.id,
             raw_part_number="832414-B21", unit_price=120.0, quantity=1, total_price=120.0,
             match_method="exact", match_status="matched", is_winner=False, fluffed_loss_price=136.62)
db.add_all([lw, ll]); db.commit()
d = Deal(bid_round_id=rnd.id, master_item_id=mw.id, winning_buyer_id=buyer.id, winning_bid_line_id=lw.id,
         part_number="H0H72108CLAR8000", description="hitachi h0h72108clar8000", quantity=55,
         winning_price=136.62, total_value=136.62*55, status="approved")
db.add(d); db.commit()

data = export_buyer_award_sheet(db, rnd.id, buyer.id)
wb = load_workbook(io.BytesIO(data))
ws = wb["Award Sheet"]
# find the won row and lost row
rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
won_r = rows.get("H0H72108CLAR8000")
lost_r = rows.get("832414-B21")
check(won_r is not None and lost_r is not None, "both rows present in award sheet")
# model qty, not 1
check(ws.cell(row=won_r, column=3).value == 55, f"won row qty = 55 (got {ws.cell(row=won_r, column=3).value})")
check(ws.cell(row=lost_r, column=3).value == 20, f"lost row qty = model 20 not 1 (got {ws.cell(row=lost_r, column=3).value})")
# USD number format on price cells
check('$' in ws.cell(row=won_r, column=4).number_format, f"Your Price USD-formatted ({ws.cell(row=won_r, column=4).number_format})")
check('$' in ws.cell(row=lost_r, column=6).number_format, "Loss Notice Price USD-formatted")
# description proper-cased
check(ws.cell(row=lost_r, column=2).value == "HPE 832414-B21 480GB SATA SSD",
      f"description cleaned: {ws.cell(row=lost_r, column=2).value}")
# PO note somewhere in the sheet
all_text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1))
check("issue a PO" in all_text and BROKERS_EMAIL in all_text, "award sheet has PO instruction for winner")

print()
if fails:
    print(f"❌ {len(fails)} FAILED:")
    for f in fails: print("  -", f)
    sys.exit(1)
print("✅ ALL FEEDBACK-2 CHECKS PASSED")
