"""
Export bid results to Excel, CSV, or ZIP.
All functions return bytes ready to stream as a file download.
"""
import io
import csv
import zipfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.deal import Deal
from app.models.bid_line import BidLine
from app.models.master_item import MasterItem
from app.models.user import User
from app.services.file_parser import device_serial_uid
from app.services.normalizer import format_part_number, normalize_description

USD_FMT = '"$"#,##0.00'  # accounting-style USD for money cells
BROKERS_EMAIL = "brokers@thinktls.com"


# ── Shared style helpers ───────────────────────────────────────────────────────

DARK_FILL   = PatternFill("solid", fgColor="0F1629")
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
WIN_FILL    = PatternFill("solid", fgColor="0A3020")
LOSS_FILL   = PatternFill("solid", fgColor="2A1010")
NEUTRAL_FILL= PatternFill("solid", fgColor="0D1825")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
WIN_FONT     = Font(name="Calibri", bold=True, color="34D399", size=10)
LOSS_FONT    = Font(name="Calibri", color="F87171", size=10)
NEUTRAL_FONT = Font(name="Calibri", color="AABBCC", size=10)

THIN = Border(
    left=Side(style="thin", color="1E3A5F"),
    right=Side(style="thin", color="1E3A5F"),
    top=Side(style="thin", color="1E3A5F"),
    bottom=Side(style="thin", color="1E3A5F"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _style_header_row(ws, num_cols: int, row: int = 1):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[row].height = 24


def _style_data_row(ws, row: int, num_cols: int, fill, font):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = THIN
        cell.alignment = LEFT


# ── Deals Excel ───────────────────────────────────────────────────────────────

def export_deals_excel(db: Session, bid_round_id: int) -> bytes:
    deals = db.query(Deal).filter(Deal.bid_round_id == bid_round_id).all()
    rows = []
    for d in deals:
        buyer = db.query(User).filter(User.id == d.winning_buyer_id).first()
        rows.append({
            "Part Number": d.part_number,
            "Description": d.description,
            "Quantity": d.quantity,
            "Winning Price": d.winning_price,
            "Total Value": d.total_value,
            "Winner Company": buyer.company_name if buyer else "",
            "Winner Email": buyer.email if buyer else "",
            "Status": d.status,
            "Razor Deal ID": d.razor_deal_id or "",
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Deals")
    return buf.getvalue()


# ── Deals CSV ─────────────────────────────────────────────────────────────────

def export_deals_csv(db: Session, bid_round_id: int) -> bytes:
    deals = db.query(Deal).filter(Deal.bid_round_id == bid_round_id).all()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[
        "Part Number", "Description", "Quantity", "Winning Price",
        "Total Value", "Winner Company", "Winner Email", "Status", "Razor Deal ID",
    ])
    writer.writeheader()
    for d in deals:
        buyer = db.query(User).filter(User.id == d.winning_buyer_id).first()
        writer.writerow({
            "Part Number": d.part_number,
            "Description": d.description,
            "Quantity": d.quantity,
            "Winning Price": d.winning_price,
            "Total Value": d.total_value,
            "Winner Company": buyer.company_name if buyer else "",
            "Winner Email": buyer.email if buyer else "",
            "Status": d.status,
            "Razor Deal ID": d.razor_deal_id or "",
        })
    return buf.getvalue().encode()


# ── Full Bid Comparison Excel ─────────────────────────────────────────────────

def export_bid_comparison_excel(db: Session, bid_round_id: int) -> bytes:
    """A professional "bid tab": ONE ROW PER MODEL, one column per buyer showing that buyer's unit
    price, then the Winner (buyer company) and the winning price. The winning buyer's cell in each
    row is highlighted green so you can see at a glance who won each line and how the others priced it.

    Batch-fetches masters, buyers and deals up front — the old version ran a master query AND a buyer
    query for every single bid line, which took minutes on a large round; this runs a fixed handful
    of queries regardless of size.
    """
    lines = (
        db.query(BidLine)
        .filter(BidLine.bid_round_id == bid_round_id, BidLine.match_status == "matched")
        .all()
    )
    master_ids = {l.master_item_id for l in lines if l.master_item_id}
    buyer_ids  = {l.buyer_id for l in lines if l.buyer_id}
    masters = {m.id: m for m in db.query(MasterItem).filter(MasterItem.id.in_(master_ids))} if master_ids else {}
    buyers  = {u.id: u for u in db.query(User).filter(User.id.in_(buyer_ids))} if buyer_ids else {}
    # Winner per model comes from the authoritative deals table (covers award-lot overrides).
    win_by_master = {
        d.master_item_id: d
        for d in db.query(Deal).filter(Deal.bid_round_id == bid_round_id)
    }

    def _bname(uid: int) -> str:
        b = buyers.get(uid)
        return (b.company_name or b.full_name or b.email) if b else f"Buyer {uid}"

    # Stable, alphabetical buyer column order.
    buyer_cols = sorted(buyer_ids, key=lambda uid: _bname(uid).lower())

    # model_id -> {buyer_id -> unit_price} (keep each buyer's most competitive, i.e. highest, quote)
    prices_by_master: dict[int, dict[int, float]] = {}
    for l in lines:
        if not l.master_item_id or l.unit_price is None:
            continue
        row = prices_by_master.setdefault(l.master_item_id, {})
        prev = row.get(l.buyer_id)
        if prev is None or l.unit_price > prev:
            row[l.buyer_id] = l.unit_price

    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Comparison"
    ws.sheet_view.showGridLines = False

    headers = ["Model", "Description", "Qty"] + [_bname(uid) for uid in buyer_cols] + ["Winner", "Winning Price"]
    widths  = [26, 44, 7] + [18] * len(buyer_cols) + [26, 15]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.alignment = CENTER
        c.border = THIN
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "D2"

    win_cell_fill = PatternFill("solid", fgColor="C6EFCE")   # green — the winning quote
    win_cell_font = Font(name="Calibri", bold=True, color="006100", size=10)
    data_font     = Font(name="Calibri", size=10)
    money         = '#,##0.00'

    # Sort models by part number for a predictable, readable sheet.
    ordered_ids = sorted(prices_by_master.keys(), key=lambda mid: (masters[mid].part_number if mid in masters else ""))
    excel_row = 2
    for mid in ordered_ids:
        master = masters.get(mid)
        deal   = win_by_master.get(mid)
        winner_uid = deal.winning_buyer_id if deal else None
        base = [
            format_part_number(master.part_number) if master else "",
            normalize_description(master.description) if master else "",
            master.quantity if master else "",
        ]
        for ci, v in enumerate(base, start=1):
            c = ws.cell(row=excel_row, column=ci, value=v)
            c.font = data_font
            c.border = THIN
            c.alignment = CENTER if ci == 3 else LEFT
        row_prices = prices_by_master.get(mid, {})
        for j, uid in enumerate(buyer_cols):
            col = 4 + j
            price = row_prices.get(uid)
            c = ws.cell(row=excel_row, column=col, value=price)
            c.border = THIN
            c.alignment = CENTER
            c.number_format = money
            if uid == winner_uid:
                c.fill = win_cell_fill
                c.font = win_cell_font
            else:
                c.font = data_font
        # Winner + winning price
        wc = ws.cell(row=excel_row, column=4 + len(buyer_cols), value=_bname(winner_uid) if winner_uid else "—")
        wc.font = win_cell_font if winner_uid else data_font
        wc.border = THIN
        wc.alignment = LEFT
        wp = ws.cell(row=excel_row, column=5 + len(buyer_cols), value=(deal.winning_price if deal else None))
        wp.font = win_cell_font if winner_uid else data_font
        wp.border = THIN
        wp.alignment = CENTER
        wp.number_format = money
        excel_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Buyer Award Sheet ─────────────────────────────────────────────────────────

def export_buyer_award_sheet(db: Session, bid_round_id: int, buyer_id: int) -> bytes:
    buyer = db.query(User).filter(User.id == buyer_id).first()
    lines = (
        db.query(BidLine)
        .filter(
            BidLine.bid_round_id == bid_round_id,
            BidLine.buyer_id == buyer_id,
            BidLine.match_status == "matched",
        )
        .order_by(BidLine.master_item_id)
        .all()
    )
    # Use deals table as authoritative win source (covers award-lot and single-deal overrides)
    won_master_ids = {
        d.master_item_id
        for d in db.query(Deal).filter(
            Deal.bid_round_id == bid_round_id,
            Deal.winning_buyer_id == buyer_id,
        ).all()
    }

    # Plain styles — no dark backgrounds, standard business Excel
    plain_header_fill = PatternFill("solid", fgColor="2E5090")
    plain_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    plain_won_font    = Font(name="Calibri", bold=True, color="217346", size=10)  # dark green text only
    plain_lost_font   = Font(name="Calibri", color="C00000", size=10)             # dark red text only
    plain_data_font   = Font(name="Calibri", size=10)
    no_fill           = PatternFill(fill_type=None)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Award Sheet"
    ws.sheet_view.showGridLines = False

    headers = ["Part Number", "Description", "Qty", "Your Price", "Result", "Loss Notice Price"]
    widths   = [22, 50, 8, 16, 10, 20]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = plain_header_fill
        cell.font = plain_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 22

    # Winner qty comes from the master (model quantity), matching what the on-screen results show.
    masters_by_id = {
        m.id: m for m in db.query(MasterItem).filter(
            MasterItem.id.in_([l.master_item_id for l in lines if l.master_item_id])
        )
    } if lines else {}

    for row_idx, line in enumerate(lines, start=2):
        master = masters_by_id.get(line.master_item_id)
        did_win = line.master_item_id in won_master_ids
        result = "WON" if did_win else "LOST"
        qty = (master.quantity if master else line.quantity) or 1
        values = [
            format_part_number(master.part_number if master else line.raw_part_number),
            normalize_description(master.description if master else (line.description or "")),
            qty,
            line.unit_price,
            result,
            line.fluffed_loss_price if not did_win else "",
        ]
        row_font = plain_won_font if did_win else plain_lost_font
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = no_fill
            cell.font = row_font if col_idx in (5, 6) else plain_data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx in (4, 6) and isinstance(val, (int, float)):
                cell.number_format = USD_FMT  # Your Price / Loss Notice Price in USD

    # Summary row
    won = len(won_master_ids)
    lost = len(lines) - won
    summary_row = len(lines) + 3
    summary_cell = ws.cell(row=summary_row, column=1, value=f"Total Won: {won}   |   Total Lost: {lost}")
    summary_cell.font = Font(name="Calibri", bold=True, size=11)

    # Purchase-order instruction for winners — they must issue a PO to close the award.
    if won:
        po_cell = ws.cell(
            row=summary_row + 2, column=1,
            value=f"Please issue a PO to {BROKERS_EMAIL} within 24 hours of the bid being awarded.",
        )
        po_cell.font = Font(name="Calibri", bold=True, size=11, color="C00000")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── ZIP of All Buyer Award Sheets ────────────────────────────────────────────

def export_all_award_sheets_zip(db: Session, bid_round_id: int) -> bytes:
    buyer_ids = {
        row.buyer_id
        for row in db.query(BidLine.buyer_id)
        .filter(BidLine.bid_round_id == bid_round_id, BidLine.match_status == "matched")
        .distinct()
        .all()
    }

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for buyer_id in buyer_ids:
            buyer = db.query(User).filter(User.id == buyer_id).first()
            if not buyer:
                continue
            sheet_bytes = export_buyer_award_sheet(db, bid_round_id, buyer_id)
            name = f"{buyer.company_name or buyer.email}_award_{bid_round_id}.xlsx".replace(" ", "_")
            zf.writestr(name, sheet_bytes)

    return buf.getvalue()


# ── Razor-Compatible Sales Order CSV ─────────────────────────────────────────

def export_razor_csv(db: Session, bid_round_id: int) -> bytes:
    deals = (
        db.query(Deal)
        .filter(Deal.bid_round_id == bid_round_id, Deal.status == "approved")
        .all()
    )
    buf = io.StringIO()
    fieldnames = [
        "RAZOR_ITEM_CODE", "DESCRIPTION", "QTY", "UNIT_PRICE",
        "TOTAL_VALUE", "CUSTOMER_CODE", "VENDOR_CODE", "DEAL_REF",
    ]
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for d in deals:
        buyer = db.query(User).filter(User.id == d.winning_buyer_id).first()
        writer.writerow({
            "RAZOR_ITEM_CODE": d.part_number,
            "DESCRIPTION": (d.description or "")[:100],
            "QTY": d.quantity,
            "UNIT_PRICE": f"{d.winning_price:.4f}",
            "TOTAL_VALUE": f"{d.total_value:.4f}",
            "CUSTOMER_CODE": buyer.company_name if buyer else "",
            "VENDOR_CODE": buyer.email if buyer else "",
            "DEAL_REF": f"THINKTLS-{bid_round_id}-{d.id}",
        })
    return buf.getvalue().encode()


def _lookup_ci(d: dict | None, *names: str) -> str:
    """Case-insensitive lookup of the first present key among `names` in a spec dict."""
    if not d:
        return ""
    lowered = {str(k).strip().lower(): v for k, v in d.items()}
    for n in names:
        v = lowered.get(n.lower())
        if v not in (None, "", "nan", "None"):
            return str(v)
    return ""


def export_razor_per_customer_zip(db: Session, bid_round_id: int) -> bytes:
    """One Razor upload workbook PER CUSTOMER (winning buyer), with ONE ROW PER PHYSICAL DEVICE
    — Model, Serial, UID, Price — the format ThinkTLS uploads into Razor after a sale.

    Bidding is consolidated by model ("qty 8 × 15-13079-02"), but Razor needs each device. So a
    won model is expanded back to its devices using the master item's unit_details (the Serial/
    UID captured at upload), every device carrying that model's winning unit price. Falls back to
    a single row when a round predates unit_details. Returns a ZIP of <Customer>_razor_<round>.xlsx.
    """
    deals = (
        db.query(Deal)
        .filter(Deal.bid_round_id == bid_round_id, Deal.status == "approved")
        .order_by(Deal.winning_buyer_id)
        .all()
    )
    masters = {
        m.id: m for m in db.query(MasterItem).filter(MasterItem.bid_round_id == bid_round_id).all()
    }
    buyer_ids = {d.winning_buyer_id for d in deals}
    buyers = {b.id: b for b in db.query(User).filter(User.id.in_(buyer_ids)).all()} if buyer_ids else {}

    by_buyer: dict[int, list[Deal]] = {}
    for d in deals:
        by_buyer.setdefault(d.winning_buyer_id, []).append(d)

    # EXACT Razor Sales Order Upload format (column names AND order matter for a seamless import):
    #   Model | MFG | Condition | SKU | Serial | UID | Price
    # Per ThinkTLS: Model must NOT include the serial; no Qty or Deal Ref columns; MFG/Condition/
    # SKU are left blank for them to map on their side.
    headers = ["Model", "MFG", "Condition", "SKU", "Serial", "UID", "Price"]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for buyer_id, buyer_deals in by_buyer.items():
            buyer = buyers.get(buyer_id)
            cust = (buyer.company_name or buyer.full_name or f"buyer{buyer_id}") if buyer else f"buyer{buyer_id}"

            wb = Workbook()
            ws = wb.active
            ws.title = "Razor Upload"
            ws.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = HEADER_FILL
            for d in buyer_deals:
                master = masters.get(d.master_item_id)
                extra = master.extra_columns if master else None
                # Base model, never the model-serial composite — computed once per deal.
                model = format_part_number(master.model_name(d.part_number or "") if master else (d.part_number or ""))
                price = round(d.winning_price, 2) if d.winning_price is not None else None

                devices = master.unit_details if (master and master.unit_details) else None
                if not devices:
                    # Old per-device round (one device per master, serial/uid in extra_columns),
                    # or a genuinely single-unit line.
                    s = _lookup_ci(extra, "Serial", "Serial Number", "Serial#")
                    u = _lookup_ci(extra, "UID", "Uid", "Unit ID", "Asset Tag", "Asset#")
                    devices = [{"Serial": s, "UID": u}] if (s or u) else [{} for _ in range(max(1, d.quantity or 1))]
                    # Defensive: an old per-device deal PN may carry the serial appended — strip it.
                    for tok in (s, u):
                        if tok and model.endswith("-" + tok):
                            model = model[: -(len(tok) + 1)]

                for dev in devices:
                    serial, uid = device_serial_uid(dev)
                    ws.append([model, "", "", "", serial, uid, price])

            widths = [26, 14, 14, 14, 20, 16, 12]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            xbuf = io.BytesIO()
            wb.save(xbuf)
            safe = "".join(c if c.isalnum() or c in "-_ " else "_" for c in cust).strip().replace(" ", "_")
            zf.writestr(f"{safe}_razor_{bid_round_id}.xlsx", xbuf.getvalue())

        if not by_buyer:
            zf.writestr(
                "NO_APPROVED_DEALS.txt",
                "No approved deals in this round yet. Approve deals first, then download again.",
            )

    return buf.getvalue()


def export_report_pack_zip(db: Session, bid_round_id: int) -> bytes:
    """One ZIP with every report for a round, so an admin grabs everything in a single click
    instead of downloading six files. Each part is wrapped so one failing export can't sink
    the whole pack — its error is written into an _errors.txt note instead."""
    parts: list[tuple[str, bytes]] = []
    errors: list[str] = []

    def _add(name: str, fn):
        try:
            parts.append((name, fn()))
        except Exception as exc:  # keep building the pack even if one report fails
            errors.append(f"{name}: {type(exc).__name__}: {exc}")

    _add(f"deals_round_{bid_round_id}.xlsx", lambda: export_deals_excel(db, bid_round_id))
    _add(f"bid_comparison_round_{bid_round_id}.xlsx", lambda: export_bid_comparison_excel(db, bid_round_id))
    _add(f"disposition_round_{bid_round_id}.xlsx", lambda: export_disposition_report(db, bid_round_id))
    _add(f"margin_round_{bid_round_id}.xlsx", lambda: export_margin_report(db, bid_round_id))
    _add(f"razor_sales_order_round_{bid_round_id}.csv", lambda: export_razor_csv(db, bid_round_id))
    _add(f"razor_per_customer_round_{bid_round_id}.zip", lambda: export_razor_per_customer_zip(db, bid_round_id))
    _add(f"all_award_sheets_round_{bid_round_id}.zip", lambda: export_all_award_sheets_zip(db, bid_round_id))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in parts:
            zf.writestr(name, data)
        if errors:
            zf.writestr("_errors.txt", "Some reports could not be generated:\n\n" + "\n".join(errors))
    return buf.getvalue()


# ── Inventory Disposition Report ─────────────────────────────────────────────

def export_disposition_report(db: Session, bid_round_id: int) -> bytes:
    """
    Shows every master line item with disposition:
    AWARDED (has approved deal), NO_BIDS (no valid bid lines), BELOW_RESERVE,
    or PENDING (matched lines but no deal yet approved).

    Batch-fetches deals/bid-lines/buyers once up front instead of querying per master
    item — at 1818 master items the previous per-row query pattern (3 queries each,
    5400+ total) took 14s; this version runs a fixed handful of queries regardless of
    master item count.
    """
    masters = db.query(MasterItem).filter(MasterItem.bid_round_id == bid_round_id).all()

    deals_by_item = {
        d.master_item_id: d
        for d in db.query(Deal).filter(Deal.bid_round_id == bid_round_id, Deal.status == "approved")
    }
    matched_lines_by_item: dict[int, list[BidLine]] = {}
    below_reserve_by_item: set[int] = set()
    for l in db.query(BidLine).filter(BidLine.bid_round_id == bid_round_id):
        if l.match_status == "matched":
            matched_lines_by_item.setdefault(l.master_item_id, []).append(l)
        elif l.exception_type == "below_reserve":
            below_reserve_by_item.add(l.master_item_id)

    buyer_ids = {d.winning_buyer_id for d in deals_by_item.values()}
    buyers_by_id = {u.id: u for u in db.query(User).filter(User.id.in_(buyer_ids))} if buyer_ids else {}

    rows = []
    for m in masters:
        deal = deals_by_item.get(m.id)
        valid_lines = matched_lines_by_item.get(m.id, [])

        if deal:
            disposition = "AWARDED"
            winner = buyers_by_id.get(deal.winning_buyer_id)
            awarded_to = winner.company_name if winner else ""
            awarded_price = deal.winning_price
        elif valid_lines:
            disposition = "PENDING"
            awarded_to = ""
            awarded_price = None
        elif m.id in below_reserve_by_item:
            disposition = "BELOW_RESERVE"
            awarded_to = ""
            awarded_price = None
        else:
            disposition = "NO_BIDS"
            awarded_to = ""
            awarded_price = None

        rows.append({
            "Part Number": m.part_number,
            "Description": m.description,
            "Quantity Requested": m.quantity,
            "Reserve Price": m.reserve_price,
            "Bids Received": len(valid_lines),
            "Disposition": disposition,
            "Awarded To": awarded_to,
            "Awarded Price": awarded_price,
            "Total Award Value": round(awarded_price * (m.quantity or 1), 4) if awarded_price else None,
        })

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Disposition")
    return buf.getvalue()


# ── Margin Report Excel ───────────────────────────────────────────────────────

def export_margin_report(db: Session, bid_round_id: int) -> bytes:
    deals = db.query(Deal).filter(Deal.bid_round_id == bid_round_id).all()

    buyer_ids = {d.winning_buyer_id for d in deals if d.winning_buyer_id}
    buyers_by_id = {u.id: u for u in db.query(User).filter(User.id.in_(buyer_ids))} if buyer_ids else {}
    master_ids = {d.master_item_id for d in deals if d.master_item_id}
    masters_by_id = {m.id: m for m in db.query(MasterItem).filter(MasterItem.id.in_(master_ids))} if master_ids else {}

    rows = []
    for d in deals:
        buyer  = buyers_by_id.get(d.winning_buyer_id)
        master = masters_by_id.get(d.master_item_id)
        reserve = master.reserve_price if master else None
        margin  = round(d.winning_price - reserve, 4) if reserve else None
        margin_pct = round((margin / reserve * 100), 2) if (margin is not None and reserve) else None
        rows.append({
            "Part Number": d.part_number,
            "Description": d.description,
            "Quantity": d.quantity,
            "Reserve Price": reserve,
            "Winning Price": d.winning_price,
            "Margin $": margin,
            "Margin %": margin_pct,
            "Total Value": d.total_value,
            "Winner Company": buyer.company_name if buyer else "",
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Margin Report")
    return buf.getvalue()


# ── ERP Line-Item Report ───────────────────────────────────────────────────────

def export_erp_line_report(db: Session, bid_round_id: int) -> bytes:
    """
    Consolidated report for broker ERP upload.
    One row per unit (quantity × deal line), suitable for bulk sales-order import.
    Columns: Line #, Part Number, Description, Serial # (blank — fill manually),
             Unit Price, Qty, Total Value, Winning Buyer, Buyer Email, Deal ID.
    """
    from app.models.bid_round import BidRound
    round_ = db.query(BidRound).filter(BidRound.id == bid_round_id).first()
    deals  = (
        db.query(Deal)
        .filter(Deal.bid_round_id == bid_round_id, Deal.status == "approved")
        .order_by(Deal.id)
        .all()
    )
    buyer_ids = {d.winning_buyer_id for d in deals if d.winning_buyer_id}
    buyers    = {u.id: u for u in db.query(User).filter(User.id.in_(buyer_ids)).all()}

    wb  = Workbook()
    ws  = wb.active
    ws.title = "ERP Upload"
    ws.sheet_view.showGridLines = False

    # ── Header ────────────────────────────────────────────────────
    _HDR = PatternFill("solid", fgColor="1F497D")
    _HDR_F = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    _THIN_S = Side(style="thin", color="BFBFBF")
    _BORDER = Border(left=_THIN_S, right=_THIN_S, top=_THIN_S, bottom=_THIN_S)
    _CENTER = Alignment(horizontal="center", vertical="center")
    _LEFT   = Alignment(horizontal="left",   vertical="center")

    headers    = ["Line #", "Part Number", "Description", "Serial # (fill in)", "Unit Price ($)", "Qty", "Total Value ($)", "Winning Buyer", "Buyer Email", "Deal ID"]
    col_widths = [8,        22,            40,            22,                    15,               6,     16,               24,              28,            8]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _HDR_F; c.fill = _HDR; c.alignment = _CENTER; c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data — one row per unit ────────────────────────────────────
    _ALT = PatternFill("solid", fgColor="F2F6FC")
    _NRM = PatternFill("solid", fgColor="FFFFFF")
    _DAT = Font(name="Calibri", size=10)

    line_num = 1
    excel_row = 2
    for deal in deals:
        buyer = buyers.get(deal.winning_buyer_id)
        qty   = deal.quantity or 1
        for _ in range(qty):
            fill = _ALT if excel_row % 2 == 0 else _NRM
            row_vals = [
                line_num,
                deal.part_number or "",
                deal.description or "",
                "",          # Serial # — broker fills manually
                deal.winning_price,
                1,           # one unit per row for ERP bulk upload
                deal.winning_price,
                buyer.company_name if buyer else "",
                buyer.email if buyer else "",
                deal.id,
            ]
            for ci, v in enumerate(row_vals, start=1):
                c = ws.cell(row=excel_row, column=ci, value=v)
                c.fill = fill; c.font = _DAT; c.border = _BORDER
                c.alignment = _CENTER if ci in (1, 5, 6, 7, 10) else _LEFT
            ws.row_dimensions[excel_row].height = 15
            line_num  += 1
            excel_row += 1

    # ── Summary row ────────────────────────────────────────────────
    if deals:
        total_units = sum(d.quantity or 1 for d in deals)
        total_val   = sum((d.total_value or 0) for d in deals)
        _SUM_F = Font(name="Calibri", bold=True, size=10)
        _SUM_FILL = PatternFill("solid", fgColor="DCE6F1")
        for ci in range(1, len(headers) + 1):
            c = ws.cell(row=excel_row, column=ci)
            c.fill = _SUM_FILL; c.font = _SUM_F; c.border = _BORDER
        ws.cell(row=excel_row, column=1, value="TOTAL").alignment = _CENTER
        ws.cell(row=excel_row, column=6, value=total_units).alignment = _CENTER
        ws.cell(row=excel_row, column=7, value=round(total_val, 2)).alignment = _CENTER
        ws.row_dimensions[excel_row].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
