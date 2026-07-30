"""
Generate a bid template Excel file for a given round.
Clean, professional design: white/light backgrounds, thin borders, no dark fills.
- Sheet 1: Instructions (read-only)
- Sheet 2: Bid Template — everything locked except the Unit Price ($) column
Returns bytes.
"""
import io
import hashlib
from app.core.timeutil import format_et
from app.services.file_parser import device_serial_uid
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection as _Prot
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.master_item import MasterItem
from app.models.bid_round import BidRound

# Lowercase column-name fragments that identify a serial/UID column — used to
# detect unit-level rounds so the template can use the original column names
# from the admin file instead of the generic renamed headers.
_SERIAL_INDICATORS = {"uid", "serial", "s/n", "sn", "asset tag", "asset#", "barcode"}

# ── Colour palette (clean / professional) ────────────────────────────────────
_WHITE    = "FFFFFF"
_HDR_BG   = "1F497D"   # dark blue header
_HDR_FONT = "FFFFFF"   # white header text
_LOCK_BG  = "F2F2F2"   # light gray — locked/read-only cells
_EDIT_BG  = "FFFFFF"   # white — editable cells
_ALT_BG   = "F7F9FC"   # very light blue alternate row
_INSTR_BG = "F0F4FA"   # instructions sheet background
_ACCENT   = "1F497D"   # dark blue (same as header)
_BORDER_C = "BFBFBF"   # medium gray for all borders

_thin_s      = Side(style="thin",   color=_BORDER_C)
_hdr_s       = Side(style="medium", color="1F497D")
THIN_BORDER  = Border(left=_thin_s, right=_thin_s, top=_thin_s, bottom=_thin_s)
HDR_BORDER   = Border(left=_hdr_s,  right=_hdr_s,  top=_hdr_s,  bottom=_hdr_s)

HDR_FILL      = PatternFill("solid", fgColor=_HDR_BG)
LOCK_FILL     = PatternFill("solid", fgColor=_LOCK_BG)
EDIT_FILL     = PatternFill("solid", fgColor=_EDIT_BG)
ALT_FILL      = PatternFill("solid", fgColor=_ALT_BG)
ALT_FILL_LOCK = PatternFill("solid", fgColor="EBEBEB")
INSTR_FILL    = PatternFill("solid", fgColor=_INSTR_BG)
WHITE_FILL    = PatternFill("solid", fgColor=_WHITE)

LOCKED   = _Prot(locked=True)
UNLOCKED = _Prot(locked=False)

# (round_id) → (content_fingerprint, bytes). Keyed on a hash of every item field that
# affects the rendered template — NOT just the row count. Keying on count alone returned a
# stale template whenever an admin re-uploaded an edited master file with the same number of
# rows (e.g. corrected a description, quantity, or spec value), so buyers downloaded the old data.
_template_cache: dict[int, tuple[str, bytes]] = {}


def _items_fingerprint(items) -> str:
    """Stable hash of the rendered content of every master item. Any change to a value that
    appears in the template (part number, manufacturer, description, quantity, or any spec
    column) produces a different fingerprint, so the cache self-invalidates on re-upload."""
    h = hashlib.sha256()
    for it in items:
        extra = it.extra_columns or {}
        extra_repr = "|".join(f"{k}={extra[k]}" for k in sorted(extra))
        h.update(
            f"{it.row_number}\x1f{it.part_number}\x1f{it.manufacturer}\x1f"
            f"{it.description}\x1f{it.quantity}\x1f{it.category}\x1f{extra_repr}\x1e".encode(
                "utf-8", "replace"
            )
        )
    return h.hexdigest()

HDR_FONT   = Font(name="Calibri", bold=True,  color=_HDR_FONT, size=10)
LOCK_FONT  = Font(name="Calibri", bold=False, color="595959",  size=10)
EDIT_FONT  = Font(name="Calibri", bold=False, color="000000",  size=10)
TITLE_FONT = Font(name="Calibri", bold=True,  color=_ACCENT,   size=14)
INSTR_FONT = Font(name="Calibri", bold=False, color="404040",  size=10)
LABEL_FONT = Font(name="Calibri", bold=True,  color="404040",  size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def _fmt_deadline(dt) -> str:
    if not dt:
        return "See invitation email"
    # Real Eastern time with DST — matches the app UI (was a fixed -5 "EST", an hour off in summer).
    return format_et(dt)


def generate_bid_template(db: Session, round_id: int) -> bytes:
    bid_round = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not bid_round:
        raise ValueError(f"Round {round_id} not found")

    items = (
        db.query(MasterItem)
        .filter(MasterItem.bid_round_id == round_id)
        .order_by(MasterItem.row_number)
        .all()
    )

    # Return cached bytes only if the exact item content is unchanged since last generation
    fingerprint = _items_fingerprint(items)
    cached = _template_cache.get(round_id)
    if cached and cached[0] == fingerprint:
        return cached[1]

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════
    # Sheet 1 — Instructions
    # ══════════════════════════════════════════════════════════════
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i.sheet_view.showGridLines = False
    ws_i.sheet_properties.tabColor = "1F497D"
    ws_i.column_dimensions["A"].width = 26
    ws_i.column_dimensions["B"].width = 70
    ws_i.row_dimensions[1].height = 8   # top padding

    # Fill entire sheet with a light background
    for row in range(1, 25):
        for col in range(1, 4):
            ws_i.cell(row=row, column=col).fill = INSTR_FILL

    def _i_cell(row, col, value, bold=False, label=False):
        c = ws_i.cell(row=row, column=col, value=value)
        c.fill = INSTR_FILL
        c.alignment = LEFT
        c.font = LABEL_FONT if label else INSTR_FONT
        if bold:
            c.font = Font(name="Calibri", bold=True, color="404040", size=10)
        return c

    # Title
    title = ws_i.cell(row=2, column=1, value="ThinkTLS Bid Desk — Pricing Template")
    title.font = TITLE_FONT
    title.fill = INSTR_FILL
    title.alignment = LEFT
    ws_i.row_dimensions[2].height = 28

    # Info fields
    fields = [
        ("Round",     bid_round.name),
        ("Commodity", bid_round.commodity or "—"),
        ("Customer",  bid_round.customer or "—"),
        ("Deadline",  _fmt_deadline(bid_round.submission_deadline)),
        ("Items",     f"{len(items)} line items"),
    ]
    for r_off, (label, value) in enumerate(fields, start=4):
        _i_cell(r_off, 1, label, bold=True)
        _i_cell(r_off, 2, value)
        ws_i.row_dimensions[r_off].height = 18

    # Spacer
    ws_i.row_dimensions[9].height = 6

    # How-to instructions
    instructions = [
        "How to complete this template:",
        "1.  Switch to the 'Bid Template' tab below.",
        "2.  Fill in your Unit Price ($) for each item you wish to bid on.",
        "3.  Each line is the full quantity — bids are all-or-nothing, there is no partial lot.",
        "4.  Leave Unit Price blank to opt out of a line — zero means you opt out.",
        "5.  Do NOT modify any column other than Unit Price ($).",
        "6.  The 'Device Detail' tab (if present) lists every physical unit and fills in your",
        "     price automatically per model — it is read-only, for reference only.",
        "7.  Save this file and upload it via the buyer portal before the deadline.",
        "",
        "Questions? Email brokers@thinktls.com and include the round number.",
        "",
        "CONFIDENTIAL — for authorised ThinkTLS buyers only. Do not share.",
    ]
    for r_off, line in enumerate(instructions, start=10):
        c = ws_i.cell(row=r_off, column=1, value=line)
        c.fill = INSTR_FILL
        c.font = Font(
            name="Calibri",
            bold=(line.startswith("How") or line.startswith("CONF")),
            italic=line.startswith("CONF"),
            color="1F497D" if line.startswith("How") else ("CC0000" if line.startswith("CONF") else "404040"),
            size=10,
        )
        c.alignment = LEFT
        ws_i.merge_cells(start_row=r_off, start_column=1, end_row=r_off, end_column=2)
        ws_i.row_dimensions[r_off].height = 16

    ws_i.protection.sheet = True

    # ══════════════════════════════════════════════════════════════
    # Sheet 2 — Bid Template
    # ══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Bid Template")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "217346"   # green tab

    # Discover all spec columns present across all items (union of their extra_columns keys),
    # preserving insertion order so the template column sequence is stable across uploads.
    spec_keys: list[str] = []
    seen_spec: set[str] = set()
    for item in items:
        if item.extra_columns:
            for k in item.extra_columns:
                if k not in seen_spec:
                    spec_keys.append(k)
                    seen_spec.add(k)

    # Unit-level detection: if any spec column name looks like a UID/serial column, this round
    # was uploaded as a unit-level file (one row per physical unit). In that case the original
    # column names are already in spec_keys (model title, UID, grade, etc.) so we use a minimal
    # fixed section — just Row # — and let the original columns speak for themselves.
    # Non-unit-level rounds keep the standard hardcoded headers (Part Number, Description, …).
    is_unit_level_round = any(
        any(ind in k.lower() for ind in _SERIAL_INDICATORS)
        for k in spec_keys
    )

    if is_unit_level_round:
        # Original column names carry all the data — no generic renamed fixed columns needed.
        fixed_headers = ["Row #"]
    else:
        fixed_headers = ["Row #", "Part Number", "Manufacturer", "Description", "Avail Qty"]

    # Only "Unit Price ($)" is buyer-editable. "Your Qty" was removed: a buyer takes the full
    # quantity of a line or doesn't bid on it — there is no partial quantity to enter, so the
    # column was dead input that only added confusion.
    buyer_headers = ["Unit Price ($)"]

    all_headers = fixed_headers + spec_keys + buyer_headers

    # Size every column to the width of its actual content so the file opens readable, rather
    # than to a fixed guess that clipped long values (Description / Model Title) and left short
    # ones oversized. openpyxl can't trigger Excel's runtime auto-fit, so we measure the data
    # ourselves. Sampling the first rows keeps this fast on a 9,000-row template while still
    # capturing representative widths.
    _SAMPLE = 400

    def _value_for(header: str, item) -> str:
        if header == "Row #":
            return "0000"
        if header == "Part Number":
            return item.part_number or ""
        if header == "Manufacturer":
            return item.manufacturer or ""
        if header == "Description":
            return item.description or ""
        if header == "Avail Qty":
            return str(item.quantity or "")
        return str((item.extra_columns or {}).get(header, ""))

    def _col_width(header: str, editable: bool = False) -> float:
        longest = len(str(header))
        for it in items[:_SAMPLE]:
            v = _value_for(header, it)
            if v:
                longest = max(longest, len(v))
        # +2 char padding; a little wider for the editable price column so it's obvious.
        lo, hi = (14, 18) if editable else (8, 55)
        return max(lo, min(longest + 2, hi))

    all_widths = (
        [_col_width(h) for h in fixed_headers]
        + [_col_width(k) for k in spec_keys]
        + [_col_width(h, editable=True) for h in buyer_headers]
    )

    total_cols      = len(all_headers)
    editable_start  = len(fixed_headers) + len(spec_keys) + 1  # 1-indexed
    editable        = {editable_start}   # only Unit Price ($) is editable now

    for col_idx, (hdr, width) in enumerate(zip(all_headers, all_widths), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = HDR_BORDER
        cell.protection = LOCKED
    ws.row_dimensions[1].height = 22

    # Freeze the header row
    ws.freeze_panes = "A2"

    for seq_row, item in enumerate(items, start=2):
        alt = (seq_row % 2 == 0)

        if is_unit_level_round:
            fixed_values = [seq_row - 1]   # Row # only — all data lives in spec_keys
        else:
            fixed_values = [
                seq_row - 1,
                item.part_number,
                item.manufacturer or "",
                item.description or "",
                item.quantity,
            ]
        spec_values = [
            (item.extra_columns or {}).get(k, "")
            for k in spec_keys
        ]
        buyer_values = [None]  # Unit Price ($) — buyer fills this in

        all_values = fixed_values + spec_values + buyer_values

        for col_idx, value in enumerate(all_values, start=1):
            cell = ws.cell(row=seq_row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in editable:
                cell.fill       = EDIT_FILL
                cell.font       = EDIT_FONT
                cell.protection = UNLOCKED
                cell.alignment  = CENTER
            else:
                cell.fill       = ALT_FILL_LOCK if alt else LOCK_FILL
                cell.font       = LOCK_FONT
                cell.protection = LOCKED
                cell.alignment  = LEFT if col_idx == 4 else CENTER
        ws.row_dimensions[seq_row].height = 16

    # Protect sheet so only unlocked cells are editable
    ws.protection.sheet = True
    ws.protection.password = "thinktls"
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatCells         = False
    ws.protection.formatColumns       = False
    ws.protection.formatRows          = False
    ws.protection.insertRows          = False
    ws.protection.deleteRows          = False
    ws.protection.sort                = False
    ws.protection.autoFilter          = False

    # ══════════════════════════════════════════════════════════════
    # Sheet 3 — Device Detail (only when the round is device-backed)
    # ══════════════════════════════════════════════════════════════
    # ThinkTLS bid on the consolidated model (qty locked, one Offer per model), but their Razor
    # upload needs one row per physical device with Serial/UID. So we mirror the source file's
    # two-tab shape: the buyer prices the model on "Bid Template", and this read-only "Device
    # Detail" tab lists every device and pulls that model's Offer down via a formula — so the
    # moment they price 01DE973 once, all 7 of its device rows show the price, ready for Razor.
    device_items = [it for it in items if getattr(it, "unit_details", None)]
    if device_items:
        price_col_idx = editable_start                     # 1-based column of Unit Price ($)
        price_col_letter = get_column_letter(price_col_idx)
        bid_last_row = 1 + len(items)
        # VLOOKUP table spans Part Number (col B) .. price column; return offset within it.
        vlookup_table = f"'Bid Template'!$B$2:${price_col_letter}${bid_last_row}"
        ret_index = price_col_idx - 1                      # column offset within the table (Part Number = col B)

        wsd = wb.create_sheet("Device Detail")
        wsd.sheet_view.showGridLines = False
        wsd.sheet_properties.tabColor = "1F497D"
        detail_headers = ["Model", "Manufacturer", "UID", "Serial", "Condition", "Description", "Offer ($)"]
        detail_widths  = [24, 16, 18, 20, 12, 46, 12]
        for col_idx, (hdr, width) in enumerate(zip(detail_headers, detail_widths), start=1):
            wsd.column_dimensions[get_column_letter(col_idx)].width = width
            c = wsd.cell(row=1, column=col_idx, value=hdr)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = HDR_BORDER
        wsd.row_dimensions[1].height = 22
        wsd.freeze_panes = "A2"

        drow = 2
        for it in device_items:
            model = it.part_number or ""
            for dev in (it.unit_details or []):
                serial, uid = device_serial_uid(dev)
                # Offer pulls this model's price from the Bid Template tab (blank until they bid).
                offer = (
                    f'=IFERROR(VLOOKUP(A{drow},{vlookup_table},{ret_index},FALSE),"")'
                )
                values = [model, it.manufacturer or "", uid, serial, it.category or "", it.description or "", offer]
                for col_idx, value in enumerate(values, start=1):
                    cell = wsd.cell(row=drow, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                    cell.font = LOCK_FONT
                    cell.protection = LOCKED
                    cell.alignment = LEFT if col_idx == 6 else CENTER
                wsd.row_dimensions[drow].height = 15
                drow += 1

        # Read-only sheet — everything is derived; the buyer prices on the Bid Template tab.
        wsd.protection.sheet = True
        wsd.protection.password = "thinktls"
        wsd.protection.selectLockedCells = False

    # Open directly on the Bid Template tab so the buyer sees the pricing sheet
    # immediately — not the Instructions page which was covering it on open.
    wb.active = 1   # 0 = Instructions, 1 = Bid Template

    buf = io.BytesIO()
    wb.save(buf)
    result = buf.getvalue()
    _template_cache[round_id] = (fingerprint, result)
    return result
