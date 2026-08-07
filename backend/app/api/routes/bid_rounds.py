import io
import logging
import mimetypes
import os
import shutil
import asyncio
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from typing import Optional, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.services.normalizer import format_part_number, normalize_description
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text, func, case, nullslast
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.timeutil import format_et
from app.core.executors import file_parsing_executor
from app.core.security import require_admin, get_current_user
from app.db.session import get_db, SessionLocal
from app.models.approval_override import ApprovalOverride
from app.models.bid_file import BidFile
from app.models.bid_line import BidLine
from app.models.bid_round import BidRound, round_buyers
from app.models.deal import Deal
from app.models.master_item import MasterItem
from app.models.user import User
from app.services.ai_matcher import run_ai_matching
from app.services.buyer_scorer import recalculate_buyer_scores
from app.api.routes.notifications import create_notification
from app.services.email_service import (
    send_bid_invitation, send_round_results,
    send_approval_ready_email, send_exception_alert,
)
from app.services.export_service import (
    export_deals_excel, export_deals_csv, export_bid_comparison_excel,
    export_buyer_award_sheet, export_all_award_sheets_zip,
    export_razor_csv, export_razor_per_customer_zip, export_margin_report, export_disposition_report,
    export_report_pack_zip,
    export_erp_line_report,
)
from app.services.file_parser import parse_master_file
from app.services.matcher import match_bid_lines
from app.services.template_generator import generate_bid_template
from app.services.winner_selector import select_winners

_log = logging.getLogger(__name__)

router = APIRouter(prefix="/rounds", tags=["bid_rounds"])

UPLOAD_DIR = "/tmp/thinktls_uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

_ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
_MAX_UPLOAD_BYTES = 100 * 1024 * 1024  # 100 MB
# xlsx/xls are ZIP-based (PK magic); csv has no magic bytes so we only check extension
_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"


def validate_upload(content: bytes, filename: str) -> None:
    """Raise HTTPException if the upload is not a valid bid file."""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported file type '{ext}'. Allowed: .xlsx, .xls, .csv")
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File exceeds 100 MB limit")
    if len(content) == 0:
        raise HTTPException(400, "Uploaded file is empty")
    if ext == ".xlsx" and not content.startswith(_XLSX_MAGIC):
        raise HTTPException(400, "File does not appear to be a valid Excel (.xlsx) file")
    if ext == ".xls" and not content.startswith(_XLS_MAGIC):
        raise HTTPException(400, "File does not appear to be a valid Excel (.xls) file")


class RoundCreate(BaseModel):
    name: str
    commodity: str
    customer: Optional[str] = None
    submission_deadline: Optional[datetime] = None
    notes: Optional[str] = None
    reserve_price_enabled: bool = False
    auto_approve_enabled: bool = False


class RoundOut(BaseModel):
    id: int
    name: str
    commodity: str
    customer: Optional[str] = None
    status: str
    total_line_items: int
    master_file_uploaded: bool
    submission_deadline: Optional[datetime] = None
    notes: Optional[str] = None
    reserve_price_enabled: bool = False
    auto_approve_enabled: bool = False
    created_at: Optional[datetime] = None
    master_file_uploaded_at: Optional[datetime] = None
    opened_at: Optional[datetime] = None
    closed_at: Optional[datetime] = None
    processing_started_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class BuyerAssignment(BaseModel):
    buyer_ids: List[int]


@router.post("/", response_model=RoundOut)
def create_round(req: RoundCreate, db: Session = Depends(get_db), admin=Depends(require_admin)):
    r = BidRound(**req.model_dump(), created_by_id=admin.id)
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@router.get("/", response_model=list[RoundOut])
def list_rounds(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=500),
):
    return db.query(BidRound).order_by(BidRound.created_at.desc()).offset(skip).limit(limit).all()


@router.get("/report/summary")
def report_summary(db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    Global KPI summary for the reporting dashboard.
    Returns aggregated stats across all rounds.
    """
    now = datetime.now(timezone.utc)
    thirty_days_ago = now - timedelta(days=30)

    deals_30d = db.query(Deal).filter(Deal.created_at >= thirty_days_ago, Deal.status == "approved").all()
    total_deal_value_30d = round(sum(d.total_value for d in deals_30d), 2)

    all_deals = db.query(Deal).filter(Deal.status == "approved").all()
    total_deal_value = round(sum(d.total_value for d in all_deals), 2)

    # Average margin % — pre-fetch masters once to avoid N+1
    master_ids_d = {d.master_item_id for d in all_deals}
    masters_idx = {
        m.id: m for m in db.query(MasterItem).filter(MasterItem.id.in_(master_ids_d)).all()
    }
    margins = []
    for d in all_deals:
        master = masters_idx.get(d.master_item_id)
        if master and master.reserve_price and master.reserve_price > 0:
            margins.append((d.winning_price - master.reserve_price) / master.reserve_price * 100)
    avg_margin_pct = round(sum(margins) / len(margins), 2) if margins else 0.0

    active_buyers = db.query(User).filter(User.role == "buyer", User.is_active == True).count()

    all_rounds = db.query(BidRound).all()
    total_lines = db.query(BidLine).count()
    unmatched = db.query(BidLine).filter(BidLine.match_status == "exception", BidLine.exception_type == "unmatched").count()
    unbid_rate = round(unmatched / total_lines * 100, 2) if total_lines > 0 else 0.0

    # Top buyers by total margin contribution
    top_buyers = (
        db.query(User)
        .filter(User.role == "buyer", User.total_lines_bid > 0)
        .order_by(User.total_margin_contribution.desc())
        .limit(10)
        .all()
    )

    return {
        "kpis": {
            "total_deal_value_30d": total_deal_value_30d,
            "total_deal_value_all_time": total_deal_value,
            "avg_margin_pct": avg_margin_pct,
            "active_buyers": active_buyers,
            "unbid_rate_pct": unbid_rate,
            "total_rounds": len(all_rounds),
            "total_deals": len(all_deals),
        },
        "top_buyers": [
            {
                "id": b.id,
                "full_name": b.full_name,
                "company_name": b.company_name,
                "win_rate": round(b.win_rate * 100, 1),
                "total_lines_won": b.total_lines_won,
                "total_lines_bid": b.total_lines_bid,
                "total_margin_contribution": round(b.total_margin_contribution, 2),
                "buyer_score": round(b.buyer_score, 1),
            }
            for b in top_buyers
        ],
        "recent_rounds": [
            {
                "id": r.id,
                "name": r.name,
                "status": r.status,
                "total_line_items": r.total_line_items,
                "created_at": r.created_at,
            }
            for r in sorted(all_rounds, key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True)[:5]
        ],
    }


@router.get("/report/round-trends")
def report_round_trends(limit: int = Query(12, ge=1, le=50), db: Session = Depends(get_db), _=Depends(require_admin)):
    """Per-round metrics for the most recent completed rounds, oldest→newest for charting:
    deal value, deals, avg winning price, participation (bid / invited) and exception rate.
    Lets the dashboard show round-over-round trends, not just all-time totals."""
    rounds = (
        db.query(BidRound)
        .filter(BidRound.status == "complete")
        .order_by(nullslast(BidRound.completed_at.desc()), BidRound.id.desc())
        .limit(limit)
        .all()
    )
    if not rounds:
        return {"rounds": []}
    rids = [r.id for r in rounds]

    # Batch every metric by round_id — no per-round query loop.
    deal_rows = (
        db.query(
            Deal.bid_round_id,
            func.count().label("deals"),
            func.coalesce(func.sum(Deal.total_value), 0).label("value"),
            func.coalesce(func.avg(Deal.winning_price), 0).label("avg_price"),
        )
        .filter(Deal.bid_round_id.in_(rids), Deal.status == "approved")
        .group_by(Deal.bid_round_id)
        .all()
    )
    deals_by_round = {row.bid_round_id: row for row in deal_rows}

    line_rows = (
        db.query(
            BidLine.bid_round_id,
            func.count().label("total"),
            func.count(func.distinct(BidLine.buyer_id)).label("participants"),
            func.count(case((BidLine.match_status == "exception", 1))).label("exceptions"),
        )
        .filter(BidLine.bid_round_id.in_(rids))
        .group_by(BidLine.bid_round_id)
        .all()
    )
    lines_by_round = {row.bid_round_id: row for row in line_rows}

    invited_rows = db.execute(
        text("SELECT round_id, COUNT(*) AS n FROM round_buyers WHERE round_id = ANY(:rids) GROUP BY round_id"),
        {"rids": rids},
    ).fetchall()
    invited_by_round = {row.round_id: row.n for row in invited_rows}

    out = []
    for r in rounds:
        d = deals_by_round.get(r.id)
        ln = lines_by_round.get(r.id)
        total_lines = ln.total if ln else 0
        exc = ln.exceptions if ln else 0
        out.append({
            "id": r.id,
            "name": r.name,
            "commodity": r.commodity,
            "completed_at": r.completed_at.isoformat() if r.completed_at else (r.created_at.isoformat() if r.created_at else None),
            "deals": d.deals if d else 0,
            "total_value": round(float(d.value), 2) if d else 0.0,
            "avg_price": round(float(d.avg_price), 2) if d else 0.0,
            "participants": ln.participants if ln else 0,
            "invited": invited_by_round.get(r.id, 0),
            "participation_pct": round((ln.participants / invited_by_round[r.id]) * 100, 1) if ln and invited_by_round.get(r.id) else 0.0,
            "exception_rate_pct": round(exc / total_lines * 100, 1) if total_lines else 0.0,
        })
    out.reverse()  # oldest → newest for left-to-right charting
    return {"rounds": out}


@router.get("/report/monthly-deal-value")
def report_monthly_deal_value(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Monthly approved deal value for the last 12 months."""
    now = datetime.now(timezone.utc)
    months = []
    for i in range(11, -1, -1):
        total_months = now.year * 12 + now.month - i
        year_m = (total_months - 1) // 12
        month_m = (total_months - 1) % 12 + 1
        first_day = datetime(year_m, month_m, 1, tzinfo=timezone.utc)
        last_day = datetime(year_m + 1, 1, 1, tzinfo=timezone.utc) if month_m == 12 \
                   else datetime(year_m, month_m + 1, 1, tzinfo=timezone.utc)
        total = (
            db.query(Deal)
            .filter(Deal.status == "approved", Deal.created_at >= first_day, Deal.created_at < last_day)
            .all()
        )
        months.append({
            "month": first_day.strftime("%b %Y"),
            "value": round(sum(d.total_value for d in total), 2),
            "count": len(total),
        })
    return months


@router.get("/{round_id}", response_model=RoundOut)
def get_round(round_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    return r


@router.post("/{round_id}/master-file")
async def upload_master_file(round_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_admin)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")

    content = await file.read()
    validate_upload(content, file.filename)
    try:
        # Parsing a large multi-sheet workbook is CPU-bound and can take several seconds.
        # Run it on a DEDICATED executor (not asyncio's default one) — sharing the default
        # pool with login's bcrypt thread-offload caused logins to stall for seconds behind
        # an in-progress upload (measured 13.5s vs 91ms under contention).
        rows = await asyncio.get_running_loop().run_in_executor(
            file_parsing_executor, parse_master_file, content, file.filename
        )
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Clear existing master items for this round
    db.query(MasterItem).filter(MasterItem.bid_round_id == round_id).delete()

    for row in rows:
        db.add(MasterItem(bid_round_id=round_id, **row))

    # Save original file to disk so buyers can download it later
    _upload_dir = f"/app/uploads/rounds/{round_id}/master"
    os.makedirs(_upload_dir, exist_ok=True)
    _safe_name = file.filename.replace(" ", "_")
    _disk_path = f"{_upload_dir}/{_safe_name}"
    with open(_disk_path, "wb") as fh:
        fh.write(content)

    r.master_file_uploaded = True
    r.total_line_items = len(rows)
    r.master_file_path = _disk_path
    r.master_file_uploaded_at = datetime.now(timezone.utc)
    db.commit()

    return {"message": f"Uploaded {len(rows)} line items", "total": len(rows)}


class RoundPatch(BaseModel):
    name: Optional[str] = None
    commodity: Optional[str] = None
    customer: Optional[str] = None
    notes: Optional[str] = None
    submission_deadline: Optional[datetime] = None
    reserve_price_enabled: Optional[bool] = None
    auto_approve_enabled: Optional[bool] = None


@router.patch("/{round_id}")
def patch_round(round_id: int, req: RoundPatch, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Edit mutable round fields after creation."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    for field, value in req.model_dump(exclude_none=True).items():
        setattr(r, field, value)
    db.commit()
    db.refresh(r)
    return r


@router.post("/{round_id}/open")
def open_round(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if not r.master_file_uploaded:
        raise HTTPException(400, "Upload master file before opening round")
    r.status = "open"
    r.opened_at = datetime.now(timezone.utc)
    db.commit()
    return {"status": "open"}


@router.post("/{round_id}/reopen")
def reopen_round(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Reopen a closed or errored round back to 'open' so buyers can submit again."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if r.status not in ("closed", "error"):
        raise HTTPException(400, f"Cannot reopen a round with status '{r.status}'")
    r.status = "open"
    r.opened_at = datetime.now(timezone.utc)
    db.commit()
    return {"status": "open"}


@router.post("/{round_id}/close")
def close_round(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    r.status = "closed"
    r.closed_at = datetime.now(timezone.utc)
    db.commit()
    return {"status": "closed"}


@router.get("/{round_id}/buyers")
def get_round_buyers(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    # Single JOIN query — no N+1
    rows = db.execute(
        text("""
            SELECT rb.buyer_id, rb.invite_status, rb.invited_at,
                   u.full_name, u.email, u.company_name
            FROM round_buyers rb
            JOIN users u ON u.id = rb.buyer_id
            WHERE rb.round_id = :rid
        """),
        {"rid": round_id},
    ).fetchall()
    return [
        {
            "id": row.buyer_id,
            "full_name": row.full_name,
            "email": row.email,
            "company_name": row.company_name,
            "invite_status": row.invite_status,
            "invited_at": row.invited_at,
        }
        for row in rows
    ]


@router.get("/{round_id}/participation")
def get_round_participation(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Real-time buyer participation status for a round."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")

    rows = db.execute(
        text("""
            SELECT rb.buyer_id, rb.invite_status, rb.invited_at,
                   u.full_name, u.email, u.company_name
            FROM round_buyers rb
            JOIN users u ON u.id = rb.buyer_id
            WHERE rb.round_id = :rid
        """),
        {"rid": round_id},
    ).fetchall()

    buyer_ids = [row.buyer_id for row in rows]

    # Fetch latest bid file per buyer in one query (subquery approach)
    if buyer_ids:
        subq = (
            db.query(BidFile.buyer_id, func.max(BidFile.uploaded_at).label("max_at"))
            .filter(BidFile.bid_round_id == round_id, BidFile.buyer_id.in_(buyer_ids))
            .group_by(BidFile.buyer_id)
            .subquery()
        )
        latest_files = (
            db.query(BidFile)
            .join(subq, (BidFile.buyer_id == subq.c.buyer_id) & (BidFile.uploaded_at == subq.c.max_at))
            .filter(BidFile.bid_round_id == round_id)
            .all()
        )
        bid_file_map = {bf.buyer_id: bf for bf in latest_files}

        counts_raw = (
            db.query(BidLine.buyer_id, func.count(BidLine.id).label("cnt"))
            .filter(BidLine.bid_round_id == round_id, BidLine.buyer_id.in_(buyer_ids))
            .group_by(BidLine.buyer_id)
            .all()
        )
        count_map: dict[int, int] = {bid_id: cnt for bid_id, cnt in counts_raw}
    else:
        bid_file_map = {}
        count_map = {}

    result = []
    for row in rows:
        bf = bid_file_map.get(row.buyer_id)
        result.append({
            "id": row.buyer_id,
            "full_name": row.full_name,
            "email": row.email,
            "company_name": row.company_name,
            "invite_status": row.invite_status,
            "invited_at": row.invited_at.isoformat() if row.invited_at else None,
            "uploaded_at": bf.uploaded_at.isoformat() if bf and bf.uploaded_at else None,
            "lines_submitted": count_map.get(row.buyer_id, 0) if bf else 0,
            "file_name": bf.filename if bf else None,
        })

    total = len(result)
    sent = sum(1 for r in result if r["invite_status"] in ("sent", "uploaded", "processing", "ready"))
    uploaded = sum(1 for r in result if r["uploaded_at"])
    pending = sum(1 for r in result if r["invite_status"] == "pending")

    return {
        "buyers": result,
        "stats": {
            "total": total,
            "sent": sent,
            "uploaded": uploaded,
            "pending_invite": pending,
            "no_response": sent - uploaded,
        },
    }


@router.post("/{round_id}/buyers")
def assign_buyers(round_id: int, req: BuyerAssignment, db: Session = Depends(get_db), _=Depends(require_admin)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")

    new_ids = set(req.buyer_ids)
    existing_rows = db.execute(
        text("SELECT buyer_id FROM round_buyers WHERE round_id = :rid"),
        {"rid": round_id},
    ).fetchall()
    existing_ids = {row.buyer_id for row in existing_rows}

    # Remove buyers no longer in the list — one DELETE per removed buyer (safe parameterized)
    for bid_id in existing_ids - new_ids:
        db.execute(
            text("DELETE FROM round_buyers WHERE round_id = :rid AND buyer_id = :bid"),
            {"rid": round_id, "bid": bid_id},
        )

    # Insert newly added buyers
    for bid_id in new_ids - existing_ids:
        db.execute(
            text("INSERT INTO round_buyers (round_id, buyer_id, invite_status) VALUES (:rid, :bid, 'pending')"),
            {"rid": round_id, "bid": bid_id},
        )

    db.commit()
    return {"assigned": len(req.buyer_ids)}


@router.get("/{round_id}/generate-template")
def download_template(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if not r.master_file_uploaded:
        raise HTTPException(400, "Upload master file before generating template")
    try:
        data = generate_bid_template(db, round_id)
    except Exception as exc:
        raise HTTPException(500, detail=f"Template generation failed: {type(exc).__name__}: {exc}")
    filename = f"bid_template_{round_id}.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/{round_id}/send-invitations")
def send_invitations(
    round_id: int,
    background_tasks: BackgroundTasks,
    resend: bool = Query(
        False,
        description="Send to every assigned buyer again, including ones already invited.",
    ),
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    """Email the bid invitation to assigned buyers.

    By default only buyers still marked 'pending' are contacted, so clicking twice doesn't
    spam anyone. That also meant the second click failed with an error and sent nothing —
    which read as "invitations don't work" whenever a buyer needed the mail again (lost it,
    it went to spam, or the deadline moved). `resend=true` re-sends to every assigned buyer.
    """
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if not r.master_file_uploaded:
        raise HTTPException(400, "Upload master file before sending invitations")

    if resend:
        assigned = db.execute(
            text("SELECT buyer_id FROM round_buyers WHERE round_id = :rid"), {"rid": round_id}
        ).fetchall()
        if not assigned:
            raise HTTPException(400, "No buyers are assigned to this round yet — assign buyers first.")
    else:
        assigned = db.execute(
            text("SELECT buyer_id FROM round_buyers WHERE round_id = :rid AND invite_status = 'pending'"), {"rid": round_id}
        ).fetchall()
        if not assigned:
            raise HTTPException(
                400,
                "Every assigned buyer has already been invited. Use Resend Invitations to "
                "email them again (for example if a buyer lost the mail or the deadline changed).",
            )

    # Real Eastern time (DST-aware) so the deadline in the invitation email matches what the
    # admin sees in the app — the old fixed -5 "EST" was an hour behind during daylight saving.
    if r.submission_deadline:
        deadline_str = format_et(r.submission_deadline)
    else:
        deadline_str = "See admin for deadline"
    upload_url = f"{settings.FRONTEND_URL}/portal/bid?round={round_id}"

    buyer_ids = [row.buyer_id for row in assigned]
    buyers = {b.id: b for b in db.query(User).filter(User.id.in_(buyer_ids), User.is_active == True).all()}

    sent = 0
    for row in assigned:
        buyer = buyers.get(row.buyer_id)
        if not buyer:
            continue
        background_tasks.add_task(send_bid_invitation, buyer.email, buyer.full_name, r.name, r.commodity or "", deadline_str, upload_url)
        db.execute(
            # Never downgrade a buyer who has already uploaded back to 'sent' — on a resend that
            # would wipe their progress out of the participation tracker and make it look like
            # they never bid.
            text(
                "UPDATE round_buyers "
                "SET invite_status = CASE WHEN invite_status='uploaded' THEN 'uploaded' ELSE 'sent' END, "
                "    invited_at = now() "
                "WHERE round_id=:rid AND buyer_id=:bid"
            ),
            {"rid": round_id, "bid": buyer.id},
        )
        buyer.last_invited_date = datetime.now(timezone.utc)
        sent += 1

    db.commit()
    verb = "Re-sent" if resend else "Invitations queued for"
    return {
        "sent": sent,
        "resend": resend,
        "message": f"{verb} {sent} buyer(s)" if resend else f"Invitations queued for {sent} buyer(s)",
    }


@router.post("/{round_id}/send-results")
def send_results_notifications(round_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Send bid result emails to all buyers assigned to this round."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if r.status != "complete":
        raise HTTPException(400, "Round must be complete before sending results")

    assigned = db.execute(
        text("SELECT buyer_id FROM round_buyers WHERE round_id = :rid"), {"rid": round_id}
    ).fetchall()

    # Pre-fetch all buyers and matched lines in bulk — no per-buyer queries in the loop
    buyer_ids = [row.buyer_id for row in assigned]
    buyers = {b.id: b for b in db.query(User).filter(User.id.in_(buyer_ids), User.is_active == True).all()}

    masters_idx = {m.id: m for m in db.query(MasterItem).filter(MasterItem.bid_round_id == round_id).all()}

    all_matched = db.query(BidLine).filter(
        BidLine.bid_round_id == round_id, BidLine.match_status == "matched"
    ).all()
    lines_by_buyer: dict[int, list] = defaultdict(list)
    for line in all_matched:
        lines_by_buyer[line.buyer_id].append(line)

    portal_url = f"{settings.FRONTEND_URL}/portal/results?round={round_id}"
    sent = 0
    for row in assigned:
        buyer = buyers.get(row.buyer_id)
        if not buyer:
            continue
        buyer_lines = lines_by_buyer.get(buyer.id, [])
        won = sum(1 for l in buyer_lines if l.is_winner)
        lost = len(buyer_lines) - won

        won_items, lost_items = [], []
        for line in buyer_lines:
            master = masters_idx.get(line.master_item_id) if line.master_item_id else None
            if line.is_winner:
                won_items.append({
                    "part_number": format_part_number(master.part_number if master else line.raw_part_number),
                    "description": normalize_description((master.description if master else line.description) or ""),
                    "quantity": (master.quantity if master else line.quantity) or 1,
                    "your_price": line.unit_price,
                })
            elif line.unit_price is not None and line.fluffed_loss_price is not None:
                lost_items.append({
                    "part_number": format_part_number(master.part_number if master else line.raw_part_number),
                    "description": normalize_description((master.description if master else line.description) or ""),
                    "quantity": (master.quantity if master else line.quantity) or 1,
                    "your_price": line.unit_price,
                    "winning_price": line.fluffed_loss_price,
                })

        background_tasks.add_task(
            send_round_results, buyer.email, buyer.full_name, r.name, won, lost, portal_url, won_items, lost_items
        )
        sent += 1

    return {"sent": sent, "message": f"Results queued for {sent} buyer(s)"}


@router.post("/{round_id}/process")
def process_round(round_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Match all bid lines and select winners."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if r.status != "closed":
        raise HTTPException(400, "Round must be closed before processing")

    r.status = "processing"
    r.processing_started_at = datetime.now(timezone.utc)
    db.commit()

    background_tasks.add_task(_run_processing, round_id)
    return {"message": "Processing started"}


def _auto_approve_clean_round(db, round_id: int):
    """Approve every pending deal on a clean, opt-in round and email buyers their results.
    Runs inside the processing background task (its own session), so it uses the same
    result-email builder the manual Approve-All path uses — buyers get the identical email,
    including loss detail."""
    from app.api.routes.deals import _send_results_to_all_buyers
    now = datetime.now(timezone.utc)
    deals = db.query(Deal).filter(
        Deal.bid_round_id == round_id, Deal.status == "pending_approval"
    ).all()
    for d in deals:
        d.status = "approved"
        d.approved_by = "auto-approve"
        d.approved_at = now
    db.commit()
    recalculate_buyer_scores(db, round_id)
    _log.info(f"[AutoApprove] Round {round_id}: auto-approved {len(deals)} deal(s)")
    # Reuse the same grouped result-email sender as the manual flow (opens its own session).
    _send_results_to_all_buyers(round_id)
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if r:
        create_notification(
            db,
            title=f"Round auto-approved: {r.name}",
            body=f"{len(deals)} deals were auto-approved (round had no exceptions) and buyers were emailed their results.",
            category="success",
            link=f"/admin/rounds/{round_id}/deals",
        )
        db.commit()


def _run_processing(round_id: int):
    db = SessionLocal()
    try:
        master_items = db.query(MasterItem).filter(MasterItem.bid_round_id == round_id).all()
        bid_lines = (
            db.query(BidLine)
            .filter(BidLine.bid_round_id == round_id, BidLine.match_status == "pending")
            .all()
        )
        match_bid_lines(bid_lines, master_items)
        db.commit()

        ai_summary = run_ai_matching(db, round_id)
        _log.info(f"Round {round_id} AI matching: {ai_summary}")

        select_winners(db, round_id)
        recalculate_buyer_scores(db, round_id)

        r = db.query(BidRound).filter(BidRound.id == round_id).first()
        if r:
            r.status = "complete"
            r.completed_at = datetime.now(timezone.utc)
            db.commit()

            exceptions_count = db.query(BidLine).filter(
                BidLine.bid_round_id == round_id,
                BidLine.match_status == "exception",
                BidLine.exception_resolved == False,
            ).count()
            if exceptions_count > 0:
                # Never auto-approve a round that still has anything needing a human decision.
                send_exception_alert(
                    settings.ADMIN_EMAIL, r.name, exceptions_count,
                    f"{settings.FRONTEND_URL}/admin/rounds/{round_id}/exceptions"
                )
            elif getattr(r, "auto_approve_enabled", False):
                # Opt-in automation: the round is clean and the admin pre-authorised auto-approve,
                # so approve every pending deal and send buyers their results — no manual click.
                _auto_approve_clean_round(db, round_id)
            else:
                deal_count = db.query(Deal).filter(Deal.bid_round_id == round_id).count()
                send_approval_ready_email(
                    settings.ADMIN_EMAIL, r.name, deal_count,
                    f"{settings.FRONTEND_URL}/admin/rounds/{round_id}"
                )

    except Exception as exc:
        _log.error(f"Processing failed for round {round_id}: {exc}", exc_info=True)
        try:
            r = db.query(BidRound).filter(BidRound.id == round_id).first()
            if r:
                r.status = "error"
                db.commit()
        except Exception:
            pass
        raise

    finally:
        db.close()


@router.post("/{round_id}/ai-match")
def trigger_ai_match(round_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Re-run AI fuzzy matcher on all remaining unmatched/partial-match exceptions."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    background_tasks.add_task(_run_ai_match_only, round_id)
    return {"message": "AI matching started in background"}


def _run_ai_match_only(round_id: int):
    db = SessionLocal()
    try:
        summary = run_ai_matching(db, round_id)
        _log.info(f"Manual AI match round {round_id}: {summary}")
    finally:
        db.close()


@router.get("/{round_id}/bid-files")
def list_bid_files(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """All files submitted by buyers for this round — for admin review and download."""
    files = (
        db.query(BidFile)
        .filter(BidFile.bid_round_id == round_id)
        .order_by(BidFile.uploaded_at.desc())
        .all()
    )
    buyer_ids = {f.buyer_id for f in files}
    buyers_map = {b.id: b for b in db.query(User).filter(User.id.in_(buyer_ids)).all()}
    return [
        {
            "id": f.id,
            "buyer_id": f.buyer_id,
            "buyer_name": buyers_map[f.buyer_id].full_name if f.buyer_id in buyers_map else None,
            "buyer_company": buyers_map[f.buyer_id].company_name if f.buyer_id in buyers_map else None,
            "filename": f.filename,
            "file_size_bytes": f.file_size_bytes,
            "lines_parsed": f.lines_parsed,
            "status": f.status,
            "uploaded_at": f.uploaded_at.isoformat() if f.uploaded_at else None,
            "has_file": f.file_path is not None and os.path.exists(f.file_path),
            "offer_terms": f.offer_terms,
        }
        for f in files
    ]


@router.get("/{round_id}/bid-files/{file_id}/download")
def download_bid_file(round_id: int, file_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Download the original file submitted by a buyer."""
    bf = db.query(BidFile).filter(BidFile.id == file_id, BidFile.bid_round_id == round_id).first()
    if not bf:
        raise HTTPException(404, "File not found")
    if not bf.file_path or not os.path.exists(bf.file_path):
        raise HTTPException(404, "File is no longer on disk")
    mime, _ = mimetypes.guess_type(bf.filename)
    mime = mime or "application/octet-stream"
    def _iter():
        with open(bf.file_path, "rb") as fh:
            yield from iter(lambda: fh.read(65536), b"")
    safe_filename = bf.filename.replace(" ", "_")
    return StreamingResponse(
        _iter(),
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
    )


@router.get("/{round_id}/bid-files/{file_id}/reconstruct")
def reconstruct_bid_file(round_id: int, file_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Reconstruct a downloadable Excel from the bid lines stored in DB when the original file is gone."""
    bf = db.query(BidFile).filter(BidFile.id == file_id, BidFile.bid_round_id == round_id).first()
    if not bf:
        raise HTTPException(404, "File not found")

    lines = (
        db.query(BidLine)
        .filter(BidLine.bid_file_id == file_id)
        .order_by(BidLine.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid Lines"

    headers = ["part_number", "description", "unit_price", "quantity", "match_status", "match_method", "exception_type", "is_winner"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, line in enumerate(lines, 2):
        ws.cell(row=row_idx, column=1, value=line.raw_part_number)
        ws.cell(row=row_idx, column=2, value=line.description)
        ws.cell(row=row_idx, column=3, value=float(line.unit_price) if line.unit_price else None)
        ws.cell(row=row_idx, column=4, value=line.quantity)
        ws.cell(row=row_idx, column=5, value=line.match_status)
        ws.cell(row=row_idx, column=6, value=line.match_method)
        ws.cell(row=row_idx, column=7, value=line.exception_type)
        ws.cell(row=row_idx, column=8, value="Yes" if line.is_winner else "No")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_stem = bf.filename.rsplit(".", 1)[0].replace(" ", "_")
    out_name = f"{safe_stem}_reconstructed.xlsx"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@router.delete("/{round_id}/bid-files/{file_id}")
def delete_bid_file(round_id: int, file_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Delete a buyer's submitted bid file and all its associated bid lines."""
    bf = db.query(BidFile).filter(BidFile.id == file_id, BidFile.bid_round_id == round_id).first()
    if not bf:
        raise HTTPException(404, "File not found")
    # Remove physical file if it still exists on disk
    if bf.file_path and os.path.exists(bf.file_path):
        try:
            os.remove(bf.file_path)
        except OSError:
            pass
    # Delete associated bid lines first
    db.query(BidLine).filter(BidLine.bid_file_id == file_id).delete(synchronize_session=False)
    db.delete(bf)
    db.commit()
    return {"deleted": True}


@router.delete("/{round_id}")
def delete_round(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Permanently delete a round and all its data (master items, bids, deals)."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if r.status in ("open", "processing"):
        raise HTTPException(400, "Cannot delete an active round. Close it first.")

    # Cascade delete in FK order (must respect foreign key dependencies):
    # approval_overrides.deal_id → deals.id → bid_lines.id
    bid_file_ids = [bf.id for bf in db.query(BidFile.id).filter(BidFile.bid_round_id == round_id).all()]

    # 1. ApprovalOverrides reference deals
    db.query(ApprovalOverride).filter(ApprovalOverride.bid_round_id == round_id).delete(synchronize_session=False)
    # 2. Deals reference bid_lines (winning_bid_line_id FK) — must go before BidLines
    db.query(Deal).filter(Deal.bid_round_id == round_id).delete(synchronize_session=False)
    # 3. BidLines (safe now that Deals are gone)
    if bid_file_ids:
        db.query(BidLine).filter(BidLine.bid_file_id.in_(bid_file_ids)).delete(synchronize_session=False)
        # Clean up uploaded files from disk
        for bf in db.query(BidFile).filter(BidFile.id.in_(bid_file_ids)).all():
            if bf.file_path and os.path.exists(bf.file_path):
                try:
                    os.remove(bf.file_path)
                except OSError:
                    pass
        db.query(BidFile).filter(BidFile.id.in_(bid_file_ids)).delete(synchronize_session=False)
    # 4. MasterItems, round_buyers, then the round itself
    db.query(MasterItem).filter(MasterItem.bid_round_id == round_id).delete(synchronize_session=False)
    db.execute(text("DELETE FROM round_buyers WHERE round_id = :rid"), {"rid": round_id})
    db.delete(r)
    db.commit()
    return {"deleted": True}


@router.get("/{round_id}/master-items/{master_item_id}/bids")
def get_item_bids(round_id: int, master_item_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """All bid lines for a single master item in a round — used by deal approval all-bids panel."""
    lines = (
        db.query(BidLine)
        .filter(BidLine.bid_round_id == round_id, BidLine.master_item_id == master_item_id)
        .order_by(BidLine.unit_price.desc())
        .all()
    )
    buyer_ids = {l.buyer_id for l in lines}
    buyers_map = {b.id: b for b in db.query(User).filter(User.id.in_(buyer_ids)).all()}
    return [
        {
            "bid_line_id": l.id,
            "buyer_id": l.buyer_id,
            "buyer_company": buyers_map[l.buyer_id].company_name if l.buyer_id in buyers_map else None,
            "buyer_email": buyers_map[l.buyer_id].email if l.buyer_id in buyers_map else None,
            "unit_price": l.unit_price,
            "quantity": l.quantity,
            "is_winner": l.is_winner,
            "match_status": l.match_status,
            "is_anomaly": l.is_anomaly,
            "fluffed_loss_price": l.fluffed_loss_price,
        }
        for l in lines
    ]


@router.get("/{round_id}/comparison")
def get_comparison(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    Returns all matched bid lines grouped by master item, with each buyer's price.
    Used by the comparison table UI.
    """
    master_items = db.query(MasterItem).filter(MasterItem.bid_round_id == round_id).order_by(MasterItem.row_number).all()
    lines = (
        db.query(BidLine)
        .filter(BidLine.bid_round_id == round_id, BidLine.match_status == "matched")
        .all()
    )

    # Columns cover every buyer ASSIGNED to the round, not just those with a matched line.
    # A buyer who never submitted, or who skipped this device, previously just had no cell —
    # indistinguishable from missing data. The admin must be able to see, per device, that a
    # given buyer did not quote it.
    assigned_ids = [
        row.buyer_id for row in db.execute(
            text("SELECT buyer_id FROM round_buyers WHERE round_id = :rid"), {"rid": round_id}
        ).fetchall()
    ]
    buyer_ids = {l.buyer_id for l in lines} | set(assigned_ids)
    buyer_map: dict[int, str] = {
        b.id: (b.company_name or b.full_name or str(b.id))
        for b in db.query(User).filter(User.id.in_(buyer_ids)).all()
    } if buyer_ids else {}

    by_item: dict = defaultdict(dict)
    for line in lines:
        company = buyer_map.get(line.buyer_id, str(line.buyer_id))
        by_item[line.master_item_id][company] = {
            "buyer_id": line.buyer_id,
            "unit_price": line.unit_price,
            "is_winner": line.is_winner,
            "is_anomaly": line.is_anomaly,
            "bid_line_id": line.id,
            # A submitted line with no price means the buyer returned the template but left
            # this device blank — they saw it and chose not to quote it.
            "quoted": line.unit_price is not None,
        }

    buyers = sorted(set(buyer_map.values()))

    rows = []
    for mi in master_items:
        item_bids = by_item.get(mi.id, {})
        # Spell out who did not quote THIS device: either no line at all, or a blank price.
        not_quoted_by = sorted(
            b for b in buyers
            if b not in item_bids or not item_bids[b].get("quoted")
        )
        row: dict = {
            "master_item_id": mi.id,
            "part_number": mi.part_number,
            "description": mi.description,
            "category": mi.category,
            "quantity": mi.quantity,
            "reserve_price": mi.reserve_price,
            "extra_columns": mi.extra_columns,
            "bids": item_bids,
            "not_quoted_by": not_quoted_by,
            "quoted_count": len(buyers) - len(not_quoted_by),
        }
        rows.append(row)

    # Per-buyer coverage: how many of the round's devices each buyer actually put a price on.
    total_items = len(master_items)
    coverage = []
    for company in buyers:
        quoted = sum(
            1 for r in rows
            if company in r["bids"] and r["bids"][company].get("quoted")
        )
        coverage.append({
            "buyer": company,
            "quoted": quoted,
            "not_quoted": total_items - quoted,
            "total_items": total_items,
            "quoted_pct": round(quoted / total_items * 100, 1) if total_items else 0.0,
        })

    return {"buyers": buyers, "rows": rows, "coverage": coverage, "total_items": total_items}


@router.get("/{round_id}/bid-lines")
def list_bid_lines(
    round_id: int,
    offset: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    buyer_id: Optional[int] = Query(None),
    match_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Paginated bid lines for a round with optional filters."""
    q = db.query(BidLine).filter(BidLine.bid_round_id == round_id)
    if buyer_id is not None:
        q = q.filter(BidLine.buyer_id == buyer_id)
    if match_status is not None:
        q = q.filter(BidLine.match_status == match_status)

    total = q.count()
    lines = q.order_by(BidLine.id).offset(offset).limit(limit).all()

    buyer_ids = {l.buyer_id for l in lines}
    buyers_map = {
        b.id: b for b in db.query(User).filter(User.id.in_(buyer_ids)).all()
    }

    return {
        "total": total,
        "offset": offset,
        "limit": limit,
        "items": [
            {
                "id": l.id,
                "raw_part_number": l.raw_part_number,
                "normalized_part_number": l.normalized_part_number,
                "description": l.description,
                "unit_price": l.unit_price,
                "quantity": l.quantity,
                "match_status": l.match_status,
                "match_method": l.match_method,
                "match_score": l.match_score,
                "exception_type": l.exception_type,
                "is_winner": l.is_winner,
                "is_anomaly": l.is_anomaly,
                "master_item_id": l.master_item_id,
                "buyer_id": l.buyer_id,
                "buyer_name": buyers_map[l.buyer_id].company_name if l.buyer_id in buyers_map else None,
            }
            for l in lines
        ],
    }


@router.get("/{round_id}/summary")
def round_summary(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    counts = db.query(
        func.count().label("total"),
        func.count(case((BidLine.match_status == "matched", 1))).label("matched"),
        func.count(case((BidLine.match_status == "exception", 1))).label("exceptions"),
        func.count(case((BidLine.is_winner == True, 1))).label("winners"),
    ).filter(BidLine.bid_round_id == round_id).one()

    breakdown = {
        (t or "unknown"): c
        for t, c in db.query(BidLine.exception_type, func.count())
        .filter(BidLine.bid_round_id == round_id, BidLine.match_status == "exception")
        .group_by(BidLine.exception_type)
        .all()
    }

    deal_agg = db.query(func.count().label("cnt"), func.sum(Deal.total_value).label("val")).filter(
        Deal.bid_round_id == round_id
    ).one()

    return {
        "total_bid_lines": counts.total,
        "matched": counts.matched,
        "exceptions": counts.exceptions,
        "winners": counts.winners,
        "deals": deal_agg.cnt or 0,
        "total_deal_value": round(float(deal_agg.val or 0), 2),
        "exception_breakdown": breakdown,
    }


@router.get("/{round_id}/processing-status")
def processing_status(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    Lightweight poll endpoint for the UI while a round is processing.
    Returns the current round status and progress counters so the frontend
    can show a real progress bar without re-fetching the full round detail.
    """
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")

    total = db.query(func.count(BidLine.id)).filter(BidLine.bid_round_id == round_id).scalar() or 0
    matched = db.query(func.count(BidLine.id)).filter(
        BidLine.bid_round_id == round_id, BidLine.match_status == "matched"
    ).scalar() or 0
    exceptions = db.query(func.count(BidLine.id)).filter(
        BidLine.bid_round_id == round_id, BidLine.match_status == "exception"
    ).scalar() or 0
    deals = db.query(func.count(Deal.id)).filter(Deal.bid_round_id == round_id).scalar() or 0

    # Processing has two phases that each take real time on a big round, so the bar tracks both.
    # Reporting only (matched+exceptions)/total hit 100% the moment matching committed and then
    # sat there for the whole winner-selection phase — the slowest part on a per-unit round
    # (~9,305 deals) — which read as "stuck at 100%".
    #   phase 1  matching        -> 0-50%
    #   phase 2  winners/deals   -> 50-100%
    # The phase-2 target is how many deals WILL exist: one per master item that has at least one
    # priced, matched bid.
    if total <= 0:
        pct, stage = 0.0, "Waiting for bids"
    else:
        match_frac = (matched + exceptions) / total
        if matched + exceptions < total:
            pct = round(match_frac * 50, 1)
            stage = f"Matching bid lines to your catalog ({matched + exceptions:,} of {total:,})"
        else:
            expected_deals = db.query(
                func.count(func.distinct(BidLine.master_item_id))
            ).filter(
                BidLine.bid_round_id == round_id,
                BidLine.match_status == "matched",
                BidLine.unit_price.isnot(None),
            ).scalar() or 0
            if r.status == "complete":
                pct, stage = 100.0, "Complete"
            elif expected_deals <= 0:
                pct, stage = 100.0, "No winning bids to award"
            else:
                pct = round(50 + min(deals / expected_deals, 1.0) * 50, 1)
                stage = f"Selecting winners and creating deals ({deals:,} of {expected_deals:,})"

    return {
        "status": r.status,
        "total_lines": total,
        "matched": matched,
        "exceptions": exceptions,
        "deals": deals,
        "progress_pct": pct,
        "stage": stage,
    }


@router.get("/{round_id}/export/deals.xlsx")
def export_deals_xlsx(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_deals_excel(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=deals_round_{round_id}.xlsx"})


@router.get("/{round_id}/export/deals.csv")
def export_deals_csv_route(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_deals_csv(db, round_id)
    return StreamingResponse(iter([data]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=deals_round_{round_id}.csv"})


@router.get("/{round_id}/export/comparison.xlsx")
def export_comparison(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_bid_comparison_excel(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=comparison_round_{round_id}.xlsx"})


@router.get("/{round_id}/export/award-sheet/{buyer_id}")
def export_single_award(round_id: int, buyer_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    buyer = db.query(User).filter(User.id == buyer_id).first()
    if not buyer:
        raise HTTPException(404, "Buyer not found")
    data = export_buyer_award_sheet(db, round_id, buyer_id)
    filename = f"award_{buyer.company_name or buyer.email}_{round_id}.xlsx".replace(" ", "_")
    return StreamingResponse(iter([data]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/{round_id}/export/all-awards.zip")
def export_all_awards(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_all_award_sheets_zip(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/zip", headers={"Content-Disposition": f"attachment; filename=all_award_sheets_round_{round_id}.zip"})


@router.get("/{round_id}/export/report-pack.zip")
def export_report_pack(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Everything for this round in one ZIP: deals, comparison, disposition, margin, Razor
    (round + per-customer) and all award sheets."""
    data = export_report_pack_zip(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/zip", headers={"Content-Disposition": f"attachment; filename=report_pack_round_{round_id}.zip"})


@router.get("/{round_id}/export/razor.csv")
def export_razor(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_razor_csv(db, round_id)
    return StreamingResponse(iter([data]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=razor_sales_order_round_{round_id}.csv"})


@router.get("/{round_id}/export/razor-per-customer.zip")
def export_razor_per_customer(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Per-customer Razor upload files (Model, Serial, UID, Price) — one CSV per winning buyer."""
    data = export_razor_per_customer_zip(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/zip", headers={"Content-Disposition": f"attachment; filename=razor_per_customer_round_{round_id}.zip"})


@router.get("/{round_id}/export/margin-report.xlsx")
def export_margin(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_margin_report(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=margin_report_round_{round_id}.xlsx"})


@router.get("/{round_id}/export/disposition.xlsx")
def export_disposition(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    data = export_disposition_report(db, round_id)
    return StreamingResponse(iter([data]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=disposition_report_round_{round_id}.xlsx"})


@router.get("/{round_id}/export/erp-report.xlsx")
def export_erp_report(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """ERP line-item report: one row per unit with part, serial placeholder, and winning price."""
    r = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not r:
        raise HTTPException(404, "Round not found")
    if r.status not in ("complete", "closed"):
        raise HTTPException(400, "ERP report is only available for completed rounds")
    data = export_erp_line_report(db, round_id)
    fname = f"erp_report_round_{round_id}.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/{round_id}/analytics")
def round_analytics(round_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    Rich analytics breakdown for a single bid round.
    Returns buyer performance, match quality, price distribution, anomalies, and timeline.
    """
    import statistics

    bid_round = db.query(BidRound).filter(BidRound.id == round_id).first()
    if not bid_round:
        raise HTTPException(404, "Round not found")

    masters = db.query(MasterItem).filter(MasterItem.bid_round_id == round_id).all()
    all_lines = db.query(BidLine).filter(BidLine.bid_round_id == round_id).all()
    all_deals = db.query(Deal).filter(Deal.bid_round_id == round_id).all()
    bid_files = db.query(BidFile).filter(BidFile.bid_round_id == round_id).all()

    matched = [l for l in all_lines if l.match_status == "matched"]
    exceptions = [l for l in all_lines if l.match_status == "exception"]
    winners = [l for l in all_lines if l.is_winner]
    anomalies = [l for l in all_lines if l.is_anomaly]
    approved_deals = [d for d in all_deals if d.status == "approved"]

    # Pre-group once (O(n)) so the per-buyer and per-item sections below are dict lookups, not
    # nested scans. Analytics on a per-device round (9,305 masters × 18,610 lines) previously
    # ran the price-distribution scan as ~173M comparisons in Python and took ~142s — well past
    # any HTTP timeout on the hosted tier.
    from collections import defaultdict as _dd
    matched_by_buyer: dict[int, list] = _dd(list)
    for l in matched:
        matched_by_buyer[l.buyer_id].append(l)
    approved_deals_by_buyer: dict[int, list] = _dd(list)
    for d in approved_deals:
        approved_deals_by_buyer[d.winning_buyer_id].append(d)
    bidfile_by_buyer: dict[int, object] = {}
    for bf in bid_files:
        bidfile_by_buyer.setdefault(bf.buyer_id, bf)
    lines_by_master: dict[int, list] = _dd(list)
    for l in all_lines:
        if l.unit_price and (l.match_status == "matched" or l.is_anomaly):
            lines_by_master[l.master_item_id].append(l)

    # Coverage: master items that received at least one valid matched bid
    master_ids_with_bids = {l.master_item_id for l in matched}
    coverage_pct = round(len(master_ids_with_bids) / len(masters) * 100, 1) if masters else 0.0

    # Match method breakdown
    method_counts: dict[str, int] = {}
    for l in matched:
        m = l.match_method or "unknown"
        method_counts[m] = method_counts.get(m, 0) + 1

    # Exception type breakdown
    exc_breakdown: dict[str, int] = {}
    for l in exceptions:
        t = l.exception_type or "unknown"
        exc_breakdown[t] = exc_breakdown.get(t, 0) + 1

    # Buyer performance table
    participating_buyer_ids = {l.buyer_id for l in all_lines}
    # One query for every participating buyer instead of one per buyer inside the loop. The map
    # is reused below to name which buyer's file an anomaly came from.
    buyer_map = {
        b.id: b for b in db.query(User).filter(User.id.in_(participating_buyer_ids)).all()
    } if participating_buyer_ids else {}
    buyer_rows = []
    for buyer_id in participating_buyer_ids:
        buyer = buyer_map.get(buyer_id)
        if not buyer:
            continue
        buyer_lines = matched_by_buyer.get(buyer_id, [])
        buyer_won = [l for l in buyer_lines if l.is_winner]
        buyer_deals = approved_deals_by_buyer.get(buyer_id, [])
        total_value = round(sum(d.total_value for d in buyer_deals), 2)
        bid_file = bidfile_by_buyer.get(buyer_id)
        buyer_rows.append({
            "id": buyer_id,
            "company_name": buyer.company_name or buyer.full_name,
            "email": buyer.email,
            "lines_bid": len(buyer_lines),
            "lines_won": len(buyer_won),
            "win_rate_pct": round(len(buyer_won) / len(buyer_lines) * 100, 1) if buyer_lines else 0.0,
            "total_value_awarded": total_value,
            "anomalies": len([l for l in buyer_lines if l.is_anomaly]),
            "submitted_at": bid_file.uploaded_at.isoformat() if bid_file and bid_file.uploaded_at else None,
        })
    buyer_rows.sort(key=lambda x: x["total_value_awarded"], reverse=True)

    # Price distribution per master item (for items with >= 2 bids)
    price_dist = []
    for master in masters:
        # lines_by_master already holds each item's priced matched/anomalous bids (an anomaly is
        # routed to the exception queue, so it must be included here or the ⚠ badge is dead code
        # — and the outlier is exactly what makes an item's spread worth showing).
        item_lines = lines_by_master.get(master.id, [])
        if len(item_lines) < 2:
            continue
        prices = [l.unit_price for l in item_lines]
        winner_line = next((l for l in item_lines if l.is_winner), None)
        price_dist.append({
            "part_number": master.part_number,
            "description": (master.description or "")[:60],
            "bids": len(prices),
            "min_price": min(prices),
            "max_price": max(prices),
            "median_price": round(statistics.median(prices), 4),
            "mean_price": round(statistics.mean(prices), 4),
            "spread_pct": round((max(prices) - min(prices)) / statistics.mean(prices) * 100, 1) if statistics.mean(prices) > 0 else 0.0,
            "winning_price": winner_line.unit_price if winner_line else None,
            "reserve_price": master.reserve_price,
            "has_anomaly": any(l.is_anomaly for l in item_lines),
            # Name WHOSE bid was flagged. A bare "anomaly" badge told the admin something was
            # wrong on this item but not whose file to look at, so they had no way to act on it.
            "anomaly_buyers": [
                {
                    "buyer_id": l.buyer_id,
                    "name": (
                        (buyer_map[l.buyer_id].company_name or buyer_map[l.buyer_id].full_name)
                        if l.buyer_id in buyer_map else f"Buyer {l.buyer_id}"
                    ),
                    "price": l.unit_price,
                    "resolved": bool(l.exception_resolved),
                }
                for l in item_lines if l.is_anomaly
            ],
        })
    price_dist.sort(key=lambda x: x["spread_pct"], reverse=True)

    # Submission timeline (bid files by upload time)
    timeline = []
    for bf in sorted(bid_files, key=lambda b: b.uploaded_at or datetime.min):
        buyer = db.query(User).filter(User.id == bf.buyer_id).first()
        timeline.append({
            "buyer_name": buyer.company_name if buyer else "",
            "filename": bf.filename,
            "lines": bf.lines_parsed or 0,
            "uploaded_at": bf.uploaded_at.isoformat() if bf.uploaded_at else None,
            "status": bf.status,
        })

    return {
        "round": {
            "id": bid_round.id,
            "name": bid_round.name,
            "commodity": bid_round.commodity,
            "status": bid_round.status,
            "submission_deadline": bid_round.submission_deadline.isoformat() if bid_round.submission_deadline else None,
        },
        "overview": {
            "total_master_items": len(masters),
            "items_with_bids": len(master_ids_with_bids),
            "coverage_pct": coverage_pct,
            "total_bid_lines": len(all_lines),
            "matched_lines": len(matched),
            "exception_lines": len(exceptions),
            "exception_rate_pct": round(len(exceptions) / len(all_lines) * 100, 1) if all_lines else 0.0,
            "anomaly_count": len(anomalies),
            "total_deals": len(all_deals),
            "approved_deals": len(approved_deals),
            "total_awarded_value": round(sum(d.total_value for d in approved_deals), 2),
            "buyers_participated": len(participating_buyer_ids),
        },
        "match_methods": method_counts,
        "exception_breakdown": exc_breakdown,
        "buyer_performance": buyer_rows,
        "price_distribution": price_dist[:30],  # top 30 widest spreads
        "submission_timeline": timeline,
    }
