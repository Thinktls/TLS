"""
Generate demo bid Excel files for live demo presentation.
Creates two buyer bid files for Round 4 (Q2 2026 Desktop Refresh — TechStart Inc)
that can be uploaded during the demo to show the full bid ingestion workflow.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "/tmp/thinktls_demo_files"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def make_bid_file(filename, company, buyer_name, items):
    """Create a styled Excel bid submission file that the file_parser can read.
    Headers MUST be on row 1 so pandas picks them up as column names.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid Submission"

    # Column headers on row 1 — parser-compatible names
    headers = ["Part Number", "Description", "Manufacturer", "Quantity", "Unit Price", "Total Price"]
    header_fill = PatternFill("solid", fgColor="0F3460")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Data rows start at row 2
    alt_fill_1 = PatternFill("solid", fgColor="F8F9FA")
    alt_fill_2 = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(size=10)
    price_font = Font(size=10, color="0F6E3D")

    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_sum = 0
    for row_offset, (pn, desc, mfr, qty, price) in enumerate(items):
        row = 2 + row_offset
        fill = alt_fill_1 if row_offset % 2 == 0 else alt_fill_2
        total = qty * price
        total_sum += total

        values = [pn, desc, mfr, qty, price, round(total, 2)]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            if col_idx in (5, 6):
                cell.font = price_font
                cell.number_format = "#,##0.00"
            else:
                cell.font = data_font
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    # Column widths
    col_widths = [24, 52, 14, 10, 18, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  Created: {path}  (${total_sum:,.2f} total)")
    return path


# ── Demo bid files for Round 4 ────────────────────────────────────────────────

ROUND4_ITEMS = [
    ("HP-ELITEDESK-800-G9",   "HP EliteDesk 800 G9 SFF i7-12700 16GB 512GB",      "HP",       30),
    ("DELL-OPTIPLEX-7010",    "Dell OptiPlex 7010 SFF i5-12500 16GB 256GB",        "Dell",     40),
    ("LEN-THINKCENTRE-M70Q",  "Lenovo ThinkCentre M70q Gen4 i5-13400T 8GB 256GB", "Lenovo",   25),
    ("HP-ELITEDISPLAY-E27Q",  "HP E27q G5 QHD 27-inch Monitor USB-C",             "HP",       60),
    ("DELL-27-MONITOR-P2723D","Dell 27 USB-C Hub Monitor P2723DE",                "Dell",     50),
    ("LEN-THINKPAD-DOCK-G2",  "Lenovo ThinkPad Universal Thunderbolt 4 Dock",     "Lenovo",   35),
    ("HP-USB-C-DOCK-G5",      "HP USB-C Dock G5 5TW10AA",                         "HP",       40),
    ("LOGITECH-MK850-BUNDLE", "Logitech MK850 Performance Wireless Keyboard+Mouse","Logitech", 80),
]

# Buyer 3 — Carol Wu (Global IT) — competitive pricing to win some items
CAROL_PRICES = [670.00, 522.00, 445.00, 285.00, 315.00, 183.00, 178.00, 86.50]
# Buyer 4 — David Singh (Premier IT) — slightly higher prices on some
DAVID_PRICES = [680.00, 530.00, 450.00, 292.00, 318.00, 186.00, 180.00, 88.00]

print("\n[make_demo_files] Generating demo bid Excel files...")

carol_items = [(pn, desc, mfr, qty, CAROL_PRICES[i])
               for i, (pn, desc, mfr, qty) in enumerate(ROUND4_ITEMS)]
david_items = [(pn, desc, mfr, qty, DAVID_PRICES[i])
               for i, (pn, desc, mfr, qty) in enumerate(ROUND4_ITEMS)]

make_bid_file("carol_wu_globalit_round4.xlsx",   "Global IT",    "Carol Wu",   carol_items)
make_bid_file("david_singh_premierit_round4.xlsx","Premier IT",   "David Singh",david_items)

# Also generate a master template with wrong part numbers to demo AI/fuzzy matching
FUZZY_ITEMS = [
    ("HP EliteDesk 800-G9",     "HP EliteDesk 800 G9 SFF Desktop i7",          "HP",     30, 688.00),
    ("DELL Optiplex 7010 SFF",  "Dell OptiPlex 7010 Small Form Factor",        "Dell",   40, 525.00),
    ("Lenovo M70q Gen4",        "Lenovo ThinkCentre M70q Gen 4 Mini PC",       "Lenovo", 25, 448.00),
    ("HP E27q-G5",              "HP E27q G5 27in QHD Monitor",                 "HP",     60, 288.00),
    ("P2723DE Dell Monitor",    "Dell P2723DE USB-C 27 Inch Monitor",          "Dell",   50, 316.00),
    ("Lenovo TB4 Dock",         "Lenovo ThinkPad Thunderbolt 4 Universal Dock","Lenovo", 35, 184.00),
    ("HP Dock G5 USB-C",        "HP USB-C Dock G5 Docking Station",            "HP",     40, 177.00),
    ("MK850 Logitech",          "Logitech MK850 Performance Keyboard Mouse",   "Logitech",80,  87.50),
]
make_bid_file("emma_thompson_nextech_fuzzy_round4.xlsx", "NexTech Systems", "Emma Thompson", FUZZY_ITEMS)

print(f"\n  All files saved to: {OUTPUT_DIR}/")
print("  Files for demo:")
for f in os.listdir(OUTPUT_DIR):
    size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
    print(f"    {f}  ({size/1024:.1f} KB)")
